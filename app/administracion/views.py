from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum, Q
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from .models import *
from .forms import *
from django.utils import timezone
from datetime import datetime, date
import json
from .services.libro_diario_service import LibroDiarioService
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

@login_required
def periodo_list(request):
    # Obtener parámetros de búsqueda
    search_query = request.GET.get('search', '')
    estado_filter = request.GET.get('estado', '')
    
    # Query base
    periodos = Periodo.objects.all()
    
    # Aplicar filtros
    if search_query:
        periodos = periodos.filter(nombre__icontains=search_query)
    
    if estado_filter:
        periodos = periodos.filter(estado=estado_filter)
    
    # Paginación
    paginator = Paginator(periodos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_periodos = Periodo.objects.count()
    activos = Periodo.objects.filter(estado='A').count()
    cerrados = Periodo.objects.filter(estado='C').count()
    
    context = {
        'periodos': page_obj,
        'total_periodos': total_periodos,
        'activos': activos,
        'cerrados': cerrados,
        'search_query': search_query,
        'estado_filter': estado_filter,
        'app_selected': 7,
    }
    return render(request, 'periodos/periodo_list.html', context)

@login_required
def periodo_create(request):
    if request.method == 'POST':
        form = PeriodoForm(request.POST)
        if form.is_valid():
            periodo = form.save(commit=False)
            periodo.creado_por = request.user
            periodo.modificado_por = request.user
            periodo.save()
            messages.success(request, f'Período "{periodo.nombre}" creado exitosamente.')
            return redirect('periodo_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = PeriodoForm()
    
    context = {
        'form': form,
        'title': 'Crear Período',
        'action': 'Crear',
        'app_selected': 7,
    }
    return render(request, 'periodos/periodo_form.html', context)

@login_required
def periodo_edit(request, pk):
    periodo = get_object_or_404(Periodo, pk=pk)
    
    if request.method == 'POST':
        form = PeriodoForm(request.POST, instance=periodo)
        if form.is_valid():
            periodo = form.save(commit=False)
            periodo.modificado_por = request.user
            periodo.save()
            messages.success(request, f'Período "{periodo.nombre}" actualizado exitosamente.')
            return redirect('periodo_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = PeriodoForm(instance=periodo)
    
    context = {
        'form': form,
        'periodo': periodo,
        'title': 'Editar Período',
        'action': 'Guardar Cambios',
        'app_selected': 7,
    }
    return render(request, 'periodos/periodo_form.html', context)

@login_required
def periodo_delete(request, pk):
    periodo = get_object_or_404(Periodo, pk=pk)
    
    if request.method == 'POST':
        nombre = periodo.nombre
        periodo.delete()
        messages.success(request, f'Período "{nombre}" eliminado exitosamente.')
        return redirect('periodo_list')
    
    context = {
        'periodo': periodo,
        'app_selected': 7,
    }
    return render(request, 'periodos/periodo_confirm_delete.html', context)

# ==================== VISTAS PARA EMPRESAS ====================

@login_required
def empresa_list(request):
    # Obtener parámetros de búsqueda
    search_query = request.GET.get('search', '')
    tipo_filter = request.GET.get('tipo', '')
    
    # Query base
    empresas = Empresa.objects.all()
    
    # Aplicar filtros
    if search_query:
        empresas = empresas.filter(
            Q(razon_social__icontains=search_query) |
            Q(nit__icontains=search_query) |
            Q(nombre_comercial__icontains=search_query) |
            Q(propietario__icontains=search_query)
        )
    
    if tipo_filter == 'sociedad':
        empresas = empresas.filter(es_sociedad=True)
    elif tipo_filter == 'individual':
        empresas = empresas.filter(es_sociedad=False)
    
    # Paginación
    paginator = Paginator(empresas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_empresas = Empresa.objects.count()
    sociedades = Empresa.objects.filter(es_sociedad=True).count()
    individuales = Empresa.objects.filter(es_sociedad=False).count()
    
    context = {
        'empresas': page_obj,
        'total_empresas': total_empresas,
        'sociedades': sociedades,
        'individuales': individuales,
        'search_query': search_query,
        'tipo_filter': tipo_filter,
        'app_selected': 7,
    }
    return render(request, 'administracion/empresa_list.html', context)

@login_required
def empresa_create(request):
    if request.method == 'POST':
        form = EmpresaForm(request.POST)
        if form.is_valid():
            empresa = form.save(commit=False)
            empresa.creado_por = request.user
            empresa.modificado_por = request.user
            empresa.save()
            messages.success(request, f'Empresa "{empresa.razon_social}" creada exitosamente.')
            return redirect('empresa_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = EmpresaForm()
    
    context = {
        'form': form,
        'title': 'Crear Empresa',
        'action': 'Crear',
        'app_selected': 7,
    }
    return render(request, 'administracion/empresa_form.html', context)

@login_required
def empresa_edit(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    
    if request.method == 'POST':
        form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            empresa = form.save(commit=False)
            empresa.modificado_por = request.user
            empresa.save()
            messages.success(request, f'Empresa "{empresa.razon_social}" actualizada exitosamente.')
            return redirect('empresa_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = EmpresaForm(instance=empresa)
    
    context = {
        'form': form,
        'empresa': empresa,
        'title': 'Editar Empresa',
        'action': 'Guardar Cambios',
        'app_selected': 7,
    }
    return render(request, 'administracion/empresa_form.html', context)

@login_required
def empresa_delete(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    
    if request.method == 'POST':
        razon_social = empresa.razon_social
        empresa.delete()
        messages.success(request, f'Empresa "{razon_social}" eliminada exitosamente.')
        return redirect('empresa_list')
    
    context = {
        'empresa': empresa,
        'app_selected': 7,
    }
    return render(request, 'administracion/empresa_confirm_delete.html', context)

# ==================== VISTAS PARA PROVEEDORES ====================

@login_required
def proveedor_list(request):
    # Obtener parámetros de búsqueda
    search_query = request.GET.get('search', '')
    
    # Query base
    proveedores = Proveedor.objects.all()
    
    # Aplicar filtros
    if search_query:
        proveedores = proveedores.filter(
            Q(nombre__icontains=search_query)
        )
    
    # Paginación
    paginator = Paginator(proveedores, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_proveedores = Proveedor.objects.count()
    
    context = {
        'proveedores': page_obj,
        'total_proveedores': total_proveedores,
        'search_query': search_query,
        'app_selected': 7,
    }
    return render(request, 'administracion/proveedor_list.html', context)

@login_required
def proveedor_create(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.creado_por = request.user
            proveedor.modificado_por = request.user
            proveedor.save()
            messages.success(request, f'Proveedor "{proveedor.nombre}" creado exitosamente.')
            return redirect('proveedor_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ProveedorForm()
    
    context = {
        'form': form,
        'title': 'Crear Proveedor',
        'action': 'Crear',
        'app_selected': 7,
    }
    return render(request, 'administracion/proveedor_form.html', context)

@login_required
def proveedor_edit(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.modificado_por = request.user
            proveedor.save()
            messages.success(request, f'Proveedor "{proveedor.nombre}" actualizado exitosamente.')
            return redirect('proveedor_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ProveedorForm(instance=proveedor)
    
    context = {
        'form': form,
        'proveedor': proveedor,
        'title': 'Editar Proveedor',
        'action': 'Guardar Cambios',
        'app_selected': 7,
    }
    return render(request, 'administracion/proveedor_form.html', context)

@login_required
def proveedor_delete(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        nombre = proveedor.nombre
        proveedor.delete()
        messages.success(request, f'Proveedor "{nombre}" eliminado exitosamente.')
        return redirect('proveedor_list')
    
    context = {
        'proveedor': proveedor,
        'app_selected': 7,
    }
    return render(request, 'administracion/proveedor_confirm_delete.html', context)


# ==================== VISTAS PARA GRUPOS (NOMENCLATURA) ====================

@login_required
def grupo_list(request):
    # Obtener parámetros de búsqueda
    search_query = request.GET.get('search', '')
    area_filter = request.GET.get('area', '')
    
    # Query base
    grupos = Grupo.objects.select_related('id_area_contable').all()
    
    # Aplicar filtros
    if search_query:
        grupos = grupos.filter(
            Q(nombre__icontains=search_query)
        )
    
    if area_filter:
        grupos = grupos.filter(id_area_contable_id=area_filter)
    
    # Paginación
    paginator = Paginator(grupos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_grupos = Grupo.objects.count()
    
    # Áreas para filtro
    areas = AreaContable.objects.all().order_by('nombre')
    
    context = {
        'grupos': page_obj,
        'total_grupos': total_grupos,
        'search_query': search_query,
        'area_filter': area_filter,
        'areas': areas,
        'app_selected': 7,
    }
    return render(request, 'administracion/grupo_list.html', context)

@login_required
def grupo_create(request):
    if request.method == 'POST':
        form = GrupoForm(request.POST)
        if form.is_valid():
            grupo = form.save(commit=False)
            grupo.creado_por = request.user
            grupo.modificado_por = request.user
            # Los campos ocultos se asignan automáticamente en save()
            grupo.save()
            messages.success(request, f'Grupo "{grupo.nombre}" creado exitosamente.')
            return redirect('grupo_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = GrupoForm()
    
    context = {
        'form': form,
        'title': 'Crear Grupo',
        'action': 'Crear',
        'app_selected': 7,
    }
    return render(request, 'administracion/grupo_form.html', context)

@login_required
def grupo_edit(request, pk):
    grupo = get_object_or_404(Grupo, pk=pk)
    
    if request.method == 'POST':
        form = GrupoForm(request.POST, instance=grupo)
        if form.is_valid():
            grupo = form.save(commit=False)
            grupo.modificado_por = request.user
            grupo.save()
            messages.success(request, f'Grupo "{grupo.nombre}" actualizado exitosamente.')
            return redirect('grupo_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = GrupoForm(instance=grupo)
    
    context = {
        'form': form,
        'grupo': grupo,
        'title': 'Editar Grupo',
        'action': 'Guardar Cambios',
        'app_selected': 7,
    }
    return render(request, 'administracion/grupo_form.html', context)

@login_required
def grupo_delete(request, pk):
    grupo = get_object_or_404(Grupo, pk=pk)
    
    if request.method == 'POST':
        nombre = grupo.nombre
        grupo.delete()
        messages.success(request, f'Grupo "{nombre}" eliminado exitosamente.')
        return redirect('grupo_list')
    
    context = {
        'grupo': grupo,
        'app_selected': 7,
    }
    return render(request, 'administracion/grupo_confirm_delete.html', context)

# ==================== VISTAS PARA SUBGRUPOS ====================

@login_required
def subgrupo_list(request, grupo_id):
    """
    Listado de subgrupos filtrados por grupo
    """
    grupo = get_object_or_404(Grupo, pk=grupo_id)
    search_query = request.GET.get('search', '')
    area_filter = request.GET.get('area', '')
    
    subgrupos = SubGrupo.objects.filter(id_grupo=grupo).select_related('id_area_contable')
    
    if search_query:
        subgrupos = subgrupos.filter(nombre__icontains=search_query)
    
    if area_filter:
        subgrupos = subgrupos.filter(id_area_contable_id=area_filter)
    
    paginator = Paginator(subgrupos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    areas = AreaContable.objects.all().order_by('nombre')
    
    context = {
        'subgrupos': page_obj,
        'grupo': grupo,
        'areas': areas,
        'search_query': search_query,
        'area_filter': area_filter,
        'total_subgrupos': subgrupos.count(),
        'app_selected': 7,
    }
    return render(request, 'administracion/subgrupo_list.html', context)


@login_required
def subgrupo_create(request, grupo_id):
    grupo = get_object_or_404(Grupo, pk=grupo_id)
    
    if request.method == 'POST':
        form = SubGrupoForm(request.POST)
        if form.is_valid():
            subgrupo = form.save(commit=False)
            subgrupo.id_grupo = grupo
            subgrupo.creado_por = request.user
            subgrupo.modificado_por = request.user
            subgrupo.save()
            messages.success(request, f'SubGrupo "{subgrupo.nombre}" creado exitosamente.')
            return redirect('subgrupo_list', grupo_id=grupo.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = SubGrupoForm(initial={'id_grupo': grupo})
    
    context = {
        'form': form,
        'grupo': grupo,
        'title': 'Crear SubGrupo',
        'action': 'Crear',
        'app_selected': 7,
    }
    return render(request, 'administracion/subgrupo_form.html', context)


@login_required
def subgrupo_edit(request, pk):
    subgrupo = get_object_or_404(SubGrupo, pk=pk)
    grupo = subgrupo.id_grupo
    
    if request.method == 'POST':
        form = SubGrupoForm(request.POST, instance=subgrupo)
        if form.is_valid():
            subgrupo = form.save(commit=False)
            subgrupo.modificado_por = request.user
            subgrupo.save()
            messages.success(request, f'SubGrupo "{subgrupo.nombre}" actualizado exitosamente.')
            return redirect('subgrupo_list', grupo_id=grupo.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = SubGrupoForm(instance=subgrupo)
    
    context = {
        'form': form,
        'subgrupo': subgrupo,
        'grupo': grupo,
        'title': 'Editar SubGrupo',
        'action': 'Guardar Cambios',
        'app_selected': 7,
    }
    return render(request, 'administracion/subgrupo_form.html', context)


@login_required
def subgrupo_delete(request, pk):
    subgrupo = get_object_or_404(SubGrupo, pk=pk)
    grupo_id = subgrupo.id_grupo.id
    nombre = subgrupo.nombre
    
    if request.method == 'POST':
        subgrupo.delete()
        messages.success(request, f'SubGrupo "{nombre}" eliminado exitosamente.')
        return redirect('subgrupo_list', grupo_id=grupo_id)
    
    context = {
        'subgrupo': subgrupo,
        'app_selected': 7,
    }
    return render(request, 'administracion/subgrupo_confirm_delete.html', context)


# ==================== VISTAS PARA CUENTAS ====================

@login_required
def cuenta_list(request, subgrupo_id):
    """
    Listado de cuentas filtradas por subgrupo
    """
    subgrupo = get_object_or_404(SubGrupo, pk=subgrupo_id)
    grupo = subgrupo.id_grupo
    search_query = request.GET.get('search', '')
    area_filter = request.GET.get('area', '')
    
    cuentas = Cuenta.objects.filter(id_subgrupo=subgrupo).select_related('id_area_contable')
    
    if search_query:
        cuentas = cuentas.filter(nombre__icontains=search_query)
    
    if area_filter:
        cuentas = cuentas.filter(id_area_contable_id=area_filter)
    
    paginator = Paginator(cuentas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    areas = AreaContable.objects.all().order_by('nombre')
    
    context = {
        'cuentas': page_obj,
        'subgrupo': subgrupo,
        'grupo': grupo,
        'areas': areas,
        'search_query': search_query,
        'area_filter': area_filter,
        'total_cuentas': cuentas.count(),
        'app_selected': 7,
    }
    return render(request, 'administracion/cuenta_list.html', context)


@login_required
def cuenta_create(request, subgrupo_id):
    subgrupo = get_object_or_404(SubGrupo, pk=subgrupo_id)
    grupo = subgrupo.id_grupo
    
    if request.method == 'POST':
        form = CuentaForm(request.POST)
        if form.is_valid():
            cuenta = form.save(commit=False)
            cuenta.id_subgrupo = subgrupo
            cuenta.creado_por = request.user
            cuenta.modificado_por = request.user
            cuenta.save()
            messages.success(request, f'Cuenta "{cuenta.nombre}" creada exitosamente.')
            return redirect('cuenta_list', subgrupo_id=subgrupo.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CuentaForm(initial={'id_subgrupo': subgrupo})
    
    context = {
        'form': form,
        'subgrupo': subgrupo,
        'grupo': grupo,
        'title': 'Crear Cuenta',
        'action': 'Crear',
        'app_selected': 7,
    }
    return render(request, 'administracion/cuenta_form.html', context)


@login_required
def cuenta_edit(request, pk):
    cuenta = get_object_or_404(Cuenta, pk=pk)
    subgrupo = cuenta.id_subgrupo
    grupo = subgrupo.id_grupo
    
    if request.method == 'POST':
        form = CuentaForm(request.POST, instance=cuenta)
        if form.is_valid():
            cuenta = form.save(commit=False)
            cuenta.modificado_por = request.user
            cuenta.save()
            messages.success(request, f'Cuenta "{cuenta.nombre}" actualizada exitosamente.')
            return redirect('cuenta_list', subgrupo_id=subgrupo.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CuentaForm(instance=cuenta)
    
    context = {
        'form': form,
        'cuenta': cuenta,
        'subgrupo': subgrupo,
        'grupo': grupo,
        'title': 'Editar Cuenta',
        'action': 'Guardar Cambios',
        'app_selected': 7,
    }
    return render(request, 'administracion/cuenta_form.html', context)


@login_required
def cuenta_delete(request, pk):
    cuenta = get_object_or_404(Cuenta, pk=pk)
    subgrupo_id = cuenta.id_subgrupo.id
    nombre = cuenta.nombre
    
    if request.method == 'POST':
        cuenta.delete()
        messages.success(request, f'Cuenta "{nombre}" eliminada exitosamente.')
        return redirect('cuenta_list', subgrupo_id=subgrupo_id)
    
    context = {
        'cuenta': cuenta,
        'app_selected': 7,
    }
    return render(request, 'administracion/cuenta_confirm_delete.html', context)


@login_required
def nomenclatura_listado(request):
    """
    Vista para mostrar la nomenclatura completa con formato jerárquico
    """
    grupos = Grupo.objects.all().order_by('id')
    
    # Construir la estructura jerárquica
    nomenclatura = []
    
    for grupo in grupos:
        # Agregar el grupo
        nomenclatura.append({
            'tipo': 'grupo',
            'id': grupo.id,
            'nombre': grupo.nombre,
            'nivel': 0,
            'icono': 'fa-folder',
            'color': 'navy'
        })
        
        # Obtener subgrupos de este grupo
        subgrupos = SubGrupo.objects.filter(id_grupo=grupo).order_by('id')
        
        for subgrupo in subgrupos:
            # Agregar el subgrupo
            nomenclatura.append({
                'tipo': 'subgrupo',
                'id': subgrupo.id,
                'nombre': subgrupo.nombre,
                'nivel': 1,
                'icono': 'fa-folder-open',
                'color': 'gold'
            })
            
            # Obtener cuentas de este subgrupo
            cuentas = Cuenta.objects.filter(id_subgrupo=subgrupo).order_by('id')
            
            for cuenta in cuentas:
                # Agregar la cuenta
                nomenclatura.append({
                    'tipo': 'cuenta',
                    'id': cuenta.id,
                    'nombre': cuenta.nombre,
                    'nivel': 2,
                    'icono': 'fa-file-alt',
                    'color': 'gray'
                })
    
    context = {
        'nomenclatura': nomenclatura,
        'total_grupos': grupos.count(),
        'total_subgrupos': SubGrupo.objects.count(),
        'total_cuentas': Cuenta.objects.count(),
        'app_selected': 7,
    }
    return render(request, 'administracion/nomenclatura_listado.html', context)

# ==================== VISTAS PARA SUCURSALES ====================

@login_required
def sucursal_list(request, empresa_id):
    """
    Listado de sucursales filtradas por empresa
    """
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    search_query = request.GET.get('search', '')
    
    sucursales = Sucursal.objects.filter(id_empresa=empresa)
    
    if search_query:
        sucursales = sucursales.filter(
            Q(nombre_comercial__icontains=search_query) |
            Q(nit__icontains=search_query) |
            Q(establecimiento__icontains=search_query) |
            Q(direccion__icontains=search_query)
        )
    
    paginator = Paginator(sucursales, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'sucursales': page_obj,
        'empresa': empresa,
        'search_query': search_query,
        'total_sucursales': sucursales.count(),
        'app_selected': 7,
    }
    return render(request, 'administracion/sucursal_list.html', context)


@login_required
def sucursal_create(request, empresa_id):
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    
    if request.method == 'POST':
        form = SucursalForm(request.POST)
        if form.is_valid():
            sucursal = form.save(commit=False)
            sucursal.id_empresa = empresa
            sucursal.creado_por = request.user
            sucursal.modificado_por = request.user
            sucursal.save()
            messages.success(request, f'Sucursal "{sucursal.nombre_comercial}" creada exitosamente.')
            return redirect('sucursal_list', empresa_id=empresa.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = SucursalForm(initial={'id_empresa': empresa})
        form.fields['id_empresa'].widget.attrs['disabled'] = True
    
    context = {
        'form': form,
        'empresa': empresa,
        'title': 'Crear Sucursal',
        'action': 'Crear',
        'app_selected': 7,
    }
    return render(request, 'administracion/sucursal_form.html', context)


@login_required
def sucursal_edit(request, pk):
    sucursal = get_object_or_404(Sucursal, pk=pk)
    empresa = sucursal.id_empresa
    
    if request.method == 'POST':
        form = SucursalForm(request.POST, instance=sucursal)
        if form.is_valid():
            sucursal = form.save(commit=False)
            sucursal.id_empresa = empresa  # Asegurar que no cambie la empresa
            sucursal.modificado_por = request.user
            sucursal.save()
            messages.success(request, f'Sucursal "{sucursal.nombre_comercial}" actualizada exitosamente.')
            return redirect('sucursal_list', empresa_id=empresa.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = SucursalForm(instance=sucursal)
        form.fields['id_empresa'].widget.attrs['disabled'] = True
    
    context = {
        'form': form,
        'sucursal': sucursal,
        'empresa': empresa,
        'title': 'Editar Sucursal',
        'action': 'Guardar Cambios',
        'app_selected': 7,
    }
    return render(request, 'administracion/sucursal_form.html', context)


@login_required
def sucursal_delete(request, pk):
    sucursal = get_object_or_404(Sucursal, pk=pk)
    empresa_id = sucursal.id_empresa.id
    nombre = sucursal.nombre_comercial
    
    if request.method == 'POST':
        sucursal.delete()
        messages.success(request, f'Sucursal "{nombre}" eliminada exitosamente.')
        return redirect('sucursal_list', empresa_id=empresa_id)
    
    context = {
        'sucursal': sucursal,
        'app_selected': 7,
    }
    return render(request, 'administracion/sucursal_confirm_delete.html', context)

# ==================== VISTAS PARA EMPRESA PERIODO ====================

# ==================== VISTAS PARA EMPRESA PERIODO ====================

@login_required
def empresa_periodo_list(request, empresa_id):
    """
    Listado de periodos asignados a una empresa
    """
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    search_query = request.GET.get('search', '')
    
    periodos_asignados = EmpresaPeriodo.objects.filter(
        id_empresa=empresa
    ).select_related('id_periodo').order_by('-id_periodo__fecha_inicial')  # Ordenar por fecha_inicial
    
    if search_query:
        periodos_asignados = periodos_asignados.filter(
            Q(id_periodo__nombre__icontains=search_query)
        )
    
    paginator = Paginator(periodos_asignados, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Periodos disponibles para asignar (no asignados a esta empresa)
    periodos_asignados_ids = EmpresaPeriodo.objects.filter(
        id_empresa=empresa
    ).values_list('id_periodo_id', flat=True)
    
    periodos_disponibles = Periodo.objects.exclude(
        id__in=periodos_asignados_ids
    ).order_by('-fecha_inicial')  # Ordenar por fecha_inicial
    
    context = {
        'periodos_asignados': page_obj,
        'empresa': empresa,
        'periodos_disponibles': periodos_disponibles,
        'search_query': search_query,
        'total_asignados': periodos_asignados.count(),
        'activo': periodos_asignados.filter(estatus=True).first(),
        'app_selected': 7,
    }
    return render(request, 'administracion/empresa_periodo_list.html', context)


@login_required
def empresa_periodo_create(request, empresa_id):
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    
    # Si viene un período específico desde GET
    periodo_id = request.GET.get('periodo')
    
    if request.method == 'POST':
        form = EmpresaPeriodoForm(request.POST)
        if form.is_valid():
            empresa_periodo = form.save(commit=False)
            empresa_periodo.id_empresa = empresa
            empresa_periodo.creado_por = request.user
            empresa_periodo.modificado_por = request.user
            
            # Si es el primer periodo de la empresa, activarlo automáticamente
            if not EmpresaPeriodo.objects.filter(id_empresa=empresa).exists():
                empresa_periodo.estatus = True
            
            empresa_periodo.save()
            messages.success(request, f'Período "{empresa_periodo.id_periodo.nombre}" asignado exitosamente.')
            return redirect('empresa_periodo_list', empresa_id=empresa.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        initial_data = {'id_empresa': empresa}
        if periodo_id:
            try:
                periodo = Periodo.objects.get(id=periodo_id)
                initial_data['id_periodo'] = periodo
            except Periodo.DoesNotExist:
                pass
        
        form = EmpresaPeriodoForm(initial=initial_data)
        form.fields['id_empresa'].widget.attrs['disabled'] = True
    
    context = {
        'form': form,
        'empresa': empresa,
        'title': 'Asignar Período',
        'action': 'Asignar',
        'app_selected': 7,
    }
    return render(request, 'administracion/empresa_periodo_form.html', context)

@login_required
def empresa_periodo_set_default(request, pk):
    """
    Establecer un período como predeterminado (activo) para la empresa
    """
    empresa_periodo = get_object_or_404(EmpresaPeriodo, pk=pk)
    empresa = empresa_periodo.id_empresa
    
    # Desactivar todos los periodos de esta empresa
    EmpresaPeriodo.objects.filter(id_empresa=empresa).update(estatus=False)
    
    # Activar el periodo seleccionado
    empresa_periodo.estatus = True
    empresa_periodo.modificado_por = request.user
    empresa_periodo.save()
    
    messages.success(request, f'Período "{empresa_periodo.id_periodo.nombre}" establecido como predeterminado.')
    return redirect('empresa_periodo_list', empresa_id=empresa.id)


@login_required
def empresa_periodo_delete(request, pk):
    empresa_periodo = get_object_or_404(EmpresaPeriodo, pk=pk)
    empresa_id = empresa_periodo.id_empresa.id
    periodo_nombre = empresa_periodo.id_periodo.nombre
    
    if request.method == 'POST':
        empresa_periodo.delete()
        messages.success(request, f'Período "{periodo_nombre}" desasignado exitosamente.')
        return redirect('empresa_periodo_list', empresa_id=empresa_id)
    
    context = {
        'empresa_periodo': empresa_periodo,
        'app_selected': 7,
    }
    return render(request, 'administracion/empresa_periodo_confirm_delete.html', context)


@csrf_exempt
@require_http_methods(["POST"])
def set_empresa_context(request):
    """API para establecer la empresa activa"""
    if request.method == 'POST':
        empresa_id = request.POST.get('empresa_id')
        if empresa_id:
            try:
                empresa = Empresa.objects.get(id=empresa_id)
                
                # Verificar si la empresa tiene períodos asignados
                periodos_asignados = EmpresaPeriodo.objects.filter(id_empresa=empresa)
                
                if not periodos_asignados.exists():
                    return JsonResponse({
                        'success': False,
                        'error': 'Esta empresa no tiene períodos asignados. Debe asignar un período primero.'
                    })
                
                # Buscar período activo
                periodo_activo = periodos_asignados.filter(estatus=True).first()
                
                if periodo_activo:
                    # La empresa tiene período activo - guardar el ID del EmpresaPeriodo
                    request.session['empresa_activa_id'] = int(empresa_id)
                    request.session['empresa_periodo_activo_id'] = periodo_activo.id  # Guardar ID del EmpresaPeriodo
                    
                    return JsonResponse({
                        'success': True,
                        'periodo_activo': True,
                        'periodo_nombre': periodo_activo.id_periodo.nombre,
                        'empresa_periodo_id': periodo_activo.id
                    })
                else:
                    # La empresa tiene períodos pero ninguno activo
                    periodos_list = []
                    for ep in periodos_asignados:
                        periodos_list.append({
                            'id': ep.id,
                            'periodo_id': ep.id_periodo.id,
                            'nombre': ep.id_periodo.nombre
                        })
                    
                    return JsonResponse({
                        'success': True,
                        'periodo_activo': False,
                        'tiene_periodos': True,
                        'periodos': periodos_list
                    })
                    
            except Empresa.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Empresa no encontrada'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@csrf_exempt
@require_http_methods(["POST"])
def set_periodo_context(request):
    """API para establecer el período activo"""
    if request.method == 'POST':
        empresa_periodo_id = request.POST.get('periodo_id')  # Este es el ID del EmpresaPeriodo
        if empresa_periodo_id:
            try:
                empresa_periodo = EmpresaPeriodo.objects.select_related('id_periodo', 'id_empresa').get(id=empresa_periodo_id)
                
                # Activar este período y desactivar los demás
                EmpresaPeriodo.objects.filter(
                    id_empresa=empresa_periodo.id_empresa
                ).update(estatus=False)
                
                empresa_periodo.estatus = True
                empresa_periodo.save()
                
                request.session['empresa_activa_id'] = empresa_periodo.id_empresa.id
                request.session['empresa_periodo_activo_id'] = empresa_periodo.id
                
                return JsonResponse({'success': True, 'periodo_nombre': empresa_periodo.id_periodo.nombre})
            except EmpresaPeriodo.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Período no encontrado'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# ==================== FUNCIÓN PARA GENERAR CORRELATIVO ====================

def generar_correlativo(empresa_periodo_id, fecha):
    """
    Genera el siguiente correlativo para un asiento por empresa, mes y año.
    El correlativo se reinicia cada mes por empresa.
    """
    with transaction.atomic():
        empresa_periodo = EmpresaPeriodo.objects.select_related('id_empresa').get(id=empresa_periodo_id)
        empresa = empresa_periodo.id_empresa

        correlativo_obj, _ = CorrelativoAsiento.objects.select_for_update().get_or_create(
            id_empresa=empresa,
            anio=fecha.year,
            mes=fecha.month,
            defaults={'ultimo_correlativo': 0}
        )

        correlativo_obj.ultimo_correlativo += 1
        correlativo_obj.save()

        return correlativo_obj.ultimo_correlativo, fecha.year, fecha.month


# ==================== VISTAS PARA ASIENTOS ====================

@login_required
def asiento_list(request):
    """
    Listado de asientos del período activo
    """
    empresa_activa_id = request.session.get('empresa_activa_id')
    empresa_periodo_activo_id = request.session.get('empresa_periodo_activo_id')
    
    # Verificar si hay empresa y período seleccionados
    if not empresa_activa_id or not empresa_periodo_activo_id:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo para ver los asientos.')
        context = {
            'asientos': [],
            'empresa_periodo': None,
            'search_query': '',
            'estatus_filter': '',
            'total_asientos': 0,
            'app_selected': 7,
        }
        return render(request, 'administracion/asiento_list.html', context)
    
    # Verificar que el EmpresaPeriodo existe
    try:
        empresa_periodo = EmpresaPeriodo.objects.select_related('id_empresa', 'id_periodo').get(
            id=empresa_periodo_activo_id,
            id_empresa_id=empresa_activa_id
        )
    except EmpresaPeriodo.DoesNotExist:
        messages.error(request, 'El período seleccionado ya no existe. Por favor seleccione otro.')
        request.session.pop('empresa_periodo_activo_id', None)
        context = {
            'asientos': [],
            'empresa_periodo': None,
            'search_query': '',
            'estatus_filter': '',
            'total_asientos': 0,
            'app_selected': 7,
        }
        return render(request, 'administracion/asiento_list.html', context)
    
    search_query = request.GET.get('search', '')
    estatus_filter = request.GET.get('estatus', '')
    
    asientos = Asiento.objects.filter(id_empresa_periodo=empresa_periodo)
    
    if search_query:
        asientos = asientos.filter(comentario__icontains=search_query)
    
    if estatus_filter:
        asientos = asientos.filter(estatus=estatus_filter)
    
    # Agregar totales a cada asiento
    for asiento in asientos:
        asiento.total_debe = asiento.get_total_debe()
        asiento.total_haber = asiento.get_total_haber()
    
    paginator = Paginator(asientos, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'asientos': page_obj,
        'empresa_periodo': empresa_periodo,
        'search_query': search_query,
        'estatus_filter': estatus_filter,
        'total_asientos': asientos.count(),
        'app_selected': 7,
    }
    return render(request, 'administracion/asiento_list.html', context)



def asiento_create(request):
    """
    Crear nuevo asiento (pantalla única)
    """
    empresa_activa_id = request.session.get('empresa_activa_id')
    empresa_periodo_activo_id = request.session.get('empresa_periodo_activo_id')
    
    if not empresa_activa_id or not empresa_periodo_activo_id:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Debe seleccionar una empresa y un período activo'})
        messages.error(request, 'Debe seleccionar una empresa y un período activo para crear un asiento')
        return redirect('asiento_list')
    
    try:
        empresa_periodo = EmpresaPeriodo.objects.select_related('id_empresa', 'id_periodo').get(
            id=empresa_periodo_activo_id,
            id_empresa_id=empresa_activa_id
        )
    except EmpresaPeriodo.DoesNotExist:
        request.session.pop('empresa_periodo_activo_id', None)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'El período seleccionado ya no existe'})
        messages.error(request, 'El período seleccionado ya no existe')
        return redirect('asiento_list')
    
    # Manejar solicitud AJAX
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        action = request.POST.get('action')
        
        if action == 'guardar_encabezado':
            form = AsientoForm(request.POST, empresa_periodo=empresa_periodo)
            if form.is_valid():
                try:
                    with transaction.atomic():
                        asiento = form.save(commit=False)
                        asiento.id_empresa_periodo = empresa_periodo
                        asiento.creado_por = request.user
                        asiento.modificado_por = request.user
                        
                        # Generar correlativo
                        correlativo, anio, mes = generar_correlativo(empresa_periodo.id, asiento.fecha)
                        asiento.correlativo = correlativo
                        asiento.anio = anio
                        asiento.mes = mes
                        
                        asiento.save()
                        
                        return JsonResponse({
                            'success': True, 
                            'asiento_id': asiento.id, 
                            'redirect': f'/administracion/asientos/editar/{asiento.id}/'
                        })
                except Exception as e:
                    return JsonResponse({'success': False, 'error': str(e)})
            else:
                errors = {}
                for field, error_list in form.errors.items():
                    errors[field] = error_list[0]
                return JsonResponse({'success': False, 'errors': errors})
    
    # GET request - mostrar formulario
    form = AsientoForm(empresa_periodo=empresa_periodo)
    
    # Datos para selectores jerárquicos
    grupos = Grupo.objects.all().order_by('nombre')
    subgrupos = SubGrupo.objects.all().order_by('nombre')
    cuentas = Cuenta.objects.select_related('id_subgrupo').all().order_by('nombre')
    proveedores = Proveedor.objects.all().order_by('nombre')
    
    # Convertir a JSON
    subgrupos_list = list(subgrupos.values('id', 'nombre', 'id_grupo_id'))
    cuentas_list = list(cuentas.values('id', 'nombre', 'id_subgrupo_id'))
    subgrupos_json = json.dumps(subgrupos_list)
    cuentas_json = json.dumps(cuentas_list)
    
    # DEBUG - Imprimir en consola del servidor
    print("=" * 50)
    print("DEBUG - asiento_create")
    print(f"Grupos encontrados: {grupos.count()}")
    print(f"Subgrupos encontrados: {subgrupos.count()}")
    print(f"Subgrupos JSON: {subgrupos_json[:500]}...")
    print(f"Cuentas encontradas: {cuentas.count()}")
    print(f"Cuentas JSON: {cuentas_json[:500]}...")
    print("=" * 50)
    
    context = {
        'form': form,
        'asiento': None,
        'movimientos': [],
        'grupos': grupos,
        'subgrupos_json': subgrupos_json,
        'cuentas_json': cuentas_json,
        'proveedores': proveedores,
        'total_debe': 0,
        'total_haber': 0,
        'is_balanced': False,
        'empresa_periodo': empresa_periodo,
        'title': 'Nuevo Asiento',
        'action': 'Crear',
        'app_selected': 7,
    }
    return render(request, 'administracion/asiento_form.html', context)


@login_required
def asiento_edit(request, pk):
    """
    Editar asiento (solo si está en borrador)
    """
    asiento = get_object_or_404(Asiento, pk=pk)
    
    # Verificar que el asiento pertenece al período activo actual
    empresa_periodo_activo_id = request.session.get('empresa_periodo_activo_id')
    if asiento.id_empresa_periodo.id != empresa_periodo_activo_id:
        messages.warning(request, 'Este asiento no pertenece al período activo actual.')
        return redirect('asiento_list')
    
    # Manejar solicitud AJAX
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        action = request.POST.get('action')
        
        if action == 'guardar_encabezado':
            form = AsientoForm(request.POST, instance=asiento, empresa_periodo=asiento.id_empresa_periodo)
            if form.is_valid():
                asiento = form.save(commit=False)
                asiento.modificado_por = request.user
                asiento.save()
                return JsonResponse({'success': True})
            else:
                errors = {}
                for field, error_list in form.errors.items():
                    errors[field] = error_list[0]
                return JsonResponse({'success': False, 'errors': errors})
        
        elif action == 'finalizar':
            if asiento.is_balanced():
                asiento.estatus = 1
                asiento.modificado_por = request.user
                asiento.save()
                return JsonResponse({'success': True, 'redirect': '/administracion/asientos/'})
            else:
                return JsonResponse({'success': False, 'error': 'El total Debe debe ser igual al total Haber'})
    
    # GET request - mostrar formulario
    form = AsientoForm(instance=asiento, empresa_periodo=asiento.id_empresa_periodo)
    movimientos = asiento.movimientos.select_related('id_cuenta').all()
    
    # Agregar totales y estado de detalles
    for mov in movimientos:
        mov.total_detalles = mov.get_total_detalles()
        mov.detalles_completos = mov.detalles_completos()
    
    # Datos para selectores jerárquicos
    grupos = Grupo.objects.all().order_by('nombre')
    subgrupos = SubGrupo.objects.all().order_by('nombre')
    cuentas = Cuenta.objects.select_related('id_subgrupo').all().order_by('nombre')
    proveedores = Proveedor.objects.all().order_by('nombre')
    
    # Convertir a JSON
    subgrupos_list = list(subgrupos.values('id', 'nombre', 'id_grupo_id'))
    cuentas_list = list(cuentas.values('id', 'nombre', 'id_subgrupo_id'))
    subgrupos_json = json.dumps(subgrupos_list)
    cuentas_json = json.dumps(cuentas_list)
    
    # DEBUG - Imprimir en consola del servidor
    print("=" * 50)
    print("DEBUG - asiento_edit")
    print(f"Asiento ID: {asiento.id}")
    print(f"Grupos encontrados: {grupos.count()}")
    print(f"Subgrupos encontrados: {subgrupos.count()}")
    print(f"Primer subgrupo: id={subgrupos_list[0] if subgrupos_list else 'None'}")
    print(f"Cuentas encontradas: {cuentas.count()}")
    print("=" * 50)
    
    context = {
        'form': form,
        'asiento': asiento,
        'movimientos': movimientos,
        'grupos': grupos,
        'subgrupos_json': subgrupos_json,
        'cuentas_json': cuentas_json,
        'proveedores': proveedores,
        'total_debe': asiento.get_total_debe(),
        'total_haber': asiento.get_total_haber(),
        'is_balanced': asiento.is_balanced(),
        'empresa_periodo': asiento.id_empresa_periodo,
        'title': f'Editar Asiento #{asiento.correlativo}',
        'action': 'Guardar',
        'app_selected': 7,
    }
    return render(request, 'administracion/asiento_form.html', context)



@login_required
def asiento_delete(request, pk):
    """
    Eliminar asiento (solo si está en borrador)
    """
    asiento = get_object_or_404(Asiento, pk=pk)
    
    """if asiento.estatus == 1:
        messages.error(request, 'No se puede eliminar un asiento finalizado')
        return redirect('asiento_list')"""
    
    if request.method == 'POST':
        correlativo = asiento.correlativo
        asiento.delete()
        messages.success(request, f'Asiento #{correlativo} eliminado exitosamente')
        return redirect('asiento_list')
    
    context = {
        'asiento': asiento,
        'app_selected': 7,
    }
    return render(request, 'administracion/asiento_confirm_delete.html', context)

# ==================== VISTAS API PARA MOVIMIENTOS (AJAX) ====================

@login_required
@require_http_methods(["POST"])
def movimiento_create(request):
    """
    Crear movimiento vía AJAX
    """
    asiento_id = request.POST.get('asiento_id')
    asiento = get_object_or_404(Asiento, pk=asiento_id)
    
    """if asiento.estatus == 1:
        return JsonResponse({'success': False, 'error': 'No se pueden agregar movimientos a un asiento finalizado'})"""
    
    form = MovimientoForm(request.POST)
    if form.is_valid():
        movimiento = form.save(commit=False)
        movimiento.id_asiento = asiento
        movimiento.creado_por = request.user
        movimiento.modificado_por = request.user
        movimiento.save()
        
        return JsonResponse({
            'success': True,
            'movimiento': {
                'id': movimiento.id,
                'cuenta_nombre': str(movimiento.id_cuenta),
                'cuenta_id': movimiento.id_cuenta.id,
                'monto': float(movimiento.monto),
                'tipo': movimiento.tipo_movimiento,
                'tipo_texto': movimiento.get_tipo_movimiento_display()
            }
        })
    else:
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = error_list[0]
        return JsonResponse({'success': False, 'errors': errors})


@login_required
@require_http_methods(["POST"])
def movimiento_delete(request):
    """
    Eliminar movimiento vía AJAX
    """
    movimiento_id = request.POST.get('movimiento_id')
    movimiento = get_object_or_404(Movimiento, pk=movimiento_id)
    asiento = movimiento.id_asiento
    
    """if asiento.estatus == 1:
        return JsonResponse({'success': False, 'error': 'No se pueden eliminar movimientos de un asiento finalizado'})"""
    
    movimiento.delete()
    
    return JsonResponse({
        'success': True,
        'totales': {
            'debe': float(asiento.get_total_debe()),
            'haber': float(asiento.get_total_haber())
        }
    })




# ==================== VISTAS API PARA DETALLES (AJAX) ====================

@login_required
def detalle_list(request, movimiento_id):
    """
    Obtener lista de detalles de un movimiento (vía AJAX)
    """
    movimiento = get_object_or_404(Movimiento, pk=movimiento_id)
    detalles = movimiento.detalles.all().values('id', 'nombre', 'descripcion', 'monto')
    
    return JsonResponse({
        'success': True,
        'detalles': list(detalles),
        'total_detalles': float(movimiento.get_total_detalles()),
        'monto_movimiento': float(movimiento.monto)
    })


@login_required
@require_http_methods(["POST"])
def detalle_create(request):
    """
    Crear detalle de movimiento vía AJAX
    """
    movimiento_id = request.POST.get('movimiento_id')
    movimiento = get_object_or_404(Movimiento, pk=movimiento_id)
    
    form = DetalleMovimientoForm(request.POST)
    if form.is_valid():
        detalle = form.save(commit=False)
        detalle.id_movimiento = movimiento
        detalle.creado_por = request.user
        detalle.modificado_por = request.user
        detalle.save()
        
        total_detalles = movimiento.get_total_detalles()
        detalles_completos = total_detalles == movimiento.monto
        
        return JsonResponse({
            'success': True,
            'detalle': {
                'id': detalle.id,
                'nombre': detalle.nombre,
                'descripcion': detalle.descripcion or '',
                'monto': float(detalle.monto)
            },
            'total_detalles': float(total_detalles),
            'monto_movimiento': float(movimiento.monto),
            'detalles_completos': detalles_completos
        })
    else:
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = error_list[0]
        return JsonResponse({'success': False, 'errors': errors})


@login_required
@require_http_methods(["POST"])
def detalle_delete(request):
    """
    Eliminar detalle de movimiento vía AJAX
    """
    detalle_id = request.POST.get('detalle_id')
    detalle = get_object_or_404(DetalleMovimiento, pk=detalle_id)
    movimiento = detalle.id_movimiento
    
    detalle.delete()
    
    total_detalles = movimiento.get_total_detalles()
    detalles_completos = total_detalles == movimiento.monto
    
    return JsonResponse({
        'success': True,
        'total_detalles': float(total_detalles),
        'monto_movimiento': float(movimiento.monto),
        'detalles_completos': detalles_completos
    })


@login_required
@require_http_methods(["POST"])
def detalle_edit(request):
    """
    Editar detalle de movimiento vía AJAX
    """
    detalle_id  = request.POST.get('detalle_id')
    nombre      = request.POST.get('nombre', '').strip()
    descripcion = request.POST.get('descripcion', '').strip()
    monto       = request.POST.get('monto')

    if not detalle_id or not nombre or not monto:
        return JsonResponse({'success': False, 'error': 'Faltan datos requeridos'})

    detalle   = get_object_or_404(DetalleMovimiento, pk=detalle_id)
    movimiento = detalle.id_movimiento

    try:
        monto = float(monto)
        if monto <= 0:
            return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0'})
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Monto inválido'})

    detalle.nombre      = nombre
    detalle.descripcion = descripcion or None
    detalle.monto       = monto
    detalle.modificado_por = request.user
    detalle.save()

    total_detalles   = movimiento.get_total_detalles()
    detalles_completos = total_detalles == movimiento.monto

    return JsonResponse({
        'success': True,
        'detalle': {
            'id':          detalle.id,
            'nombre':      detalle.nombre,
            'descripcion': detalle.descripcion or '',
            'monto':       float(detalle.monto),
        },
        'total_detalles':    float(total_detalles),
        'monto_movimiento':  float(movimiento.monto),
        'detalles_completos': detalles_completos,
    })

@login_required
def api_subgrupos_por_grupo(request, grupo_id):
    """API para obtener subgrupos de un grupo específico"""
    try:
        subgrupos = SubGrupo.objects.filter(id_grupo_id=grupo_id).values('id', 'nombre')
        return JsonResponse({
            'success': True,
            'subgrupos': list(subgrupos)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_cuentas_por_subgrupo(request, subgrupo_id):
    """API para obtener cuentas de un subgrupo específico"""
    try:
        cuentas = Cuenta.objects.filter(id_subgrupo_id=subgrupo_id).values('id', 'nombre')
        return JsonResponse({
            'success': True,
            'cuentas': list(cuentas)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
    
@login_required
def asiento_detail(request, pk):
    """
    Ver asiento finalizado (solo lectura)
    """
    asiento = get_object_or_404(Asiento, pk=pk)
    movimientos = asiento.movimientos.select_related('id_cuenta').all()
    
    for mov in movimientos:
        mov.total_detalles = mov.get_total_detalles()
        mov.detalles_completos = mov.detalles_completos()
    
    context = {
        'asiento': asiento,
        'movimientos': movimientos,
        'total_debe': asiento.get_total_debe(),
        'total_haber': asiento.get_total_haber(),
        'is_balanced': asiento.is_balanced(),
        'app_selected': 7,
    }
    return render(request, 'administracion/asiento_detail.html', context)

# ==================== LIBRO DIARIO ====================

@login_required
def libro_diario(request):
    """
    Reporte del Libro Diario con VAN/VIENEN
    """
    # Obtener empresa y período de la sesión
    empresa_activa_id = request.session.get('empresa_activa_id')
    empresa_periodo_activo_id = request.session.get('empresa_periodo_activo_id')
    
    if not empresa_activa_id or not empresa_periodo_activo_id:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')
    
    try:
        empresa = Empresa.objects.get(id=empresa_activa_id)
        empresa_periodo = EmpresaPeriodo.objects.select_related('id_periodo').get(id=empresa_periodo_activo_id)
        periodo = empresa_periodo.id_periodo
    except (Empresa.DoesNotExist, EmpresaPeriodo.DoesNotExist):
        messages.error(request, 'No se encontró la empresa o período seleccionado')
        return redirect('home')
    
    # Procesar filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    lineas_por_pagina = LibroDiarioService.LINEAS_POR_PAGINA
    
    # Fechas por defecto: todo el período
    if fecha_desde:
        try:
            fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        except ValueError:
            fecha_desde = periodo.fecha_inicial
    else:
        fecha_desde = periodo.fecha_inicial
    
    if fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        except ValueError:
            fecha_hasta = periodo.fecha_final
    else:
        fecha_hasta = periodo.fecha_final
    
    # Validar rango de fechas
    if fecha_desde > fecha_hasta:
        messages.error(request, 'La fecha desde no puede ser mayor a la fecha hasta')
        fecha_desde = periodo.fecha_inicial
        fecha_hasta = periodo.fecha_final
    
    # Obtener asientos finalizados en el rango de fechas
    asientos = Asiento.objects.filter(
        id_empresa_periodo=empresa_periodo,
        estatus__in=[1, True],
        fecha__range=[fecha_desde, fecha_hasta]
    ).order_by('fecha', 'correlativo')
    
    if not asientos.exists():
        messages.warning(request, 'No hay asientos finalizados en el período seleccionado')
        return render(request, 'administracion/reportes/libro_diario.html', {
            'sin_datos': True,
            'empresa': empresa,
            'periodo': periodo,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'lineas_por_pagina': lineas_por_pagina,
            'app_selected': 7,
        })
    
    # Obtener datos estructurados
    datos = LibroDiarioService.get_datos_reporte(asientos, fecha_desde, fecha_hasta, empresa.razon_social, periodo.nombre)

    # Paginar con VAN/VIENEN
    paginas = LibroDiarioService.paginar_con_van_vienen(datos, lineas_por_pagina)

    # Totales generales — solo de bloques tipo asiento
    total_debe_general = sum(b['total_debe'] for b in datos if b['tipo'] == 'asiento')
    total_haber_general = sum(b['total_haber'] for b in datos if b['tipo'] == 'asiento')
    
    context = {
        'empresa': empresa,
        'periodo': periodo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'lineas_por_pagina': lineas_por_pagina,
        'paginas': paginas,
        'total_debe_general': total_debe_general,
        'total_haber_general': total_haber_general,
        'fecha_reporte': timezone.now(),
        'sin_datos': False,
        'app_selected': 7,
    }
    
    return render(request, 'administracion/reportes/libro_diario.html', context)



def _get_libro_diario_datos(request):
    """Helper compartido entre Excel y PDF."""
    empresa_activa_id = request.session.get('empresa_activa_id')
    empresa_periodo_activo_id = request.session.get('empresa_periodo_activo_id')

    if not empresa_activa_id or not empresa_periodo_activo_id:
        return None, None, None, None, None, None, None

    try:
        empresa = Empresa.objects.get(id=empresa_activa_id)
        empresa_periodo = EmpresaPeriodo.objects.select_related('id_periodo').get(id=empresa_periodo_activo_id)
        periodo = empresa_periodo.id_periodo
    except (Empresa.DoesNotExist, EmpresaPeriodo.DoesNotExist):
        return None, None, None, None, None, None, None

    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    try:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date() if fecha_desde else periodo.fecha_inicial
    except ValueError:
        fecha_desde = periodo.fecha_inicial

    try:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date() if fecha_hasta else periodo.fecha_final
    except ValueError:
        fecha_hasta = periodo.fecha_final

    asientos = Asiento.objects.filter(
        id_empresa_periodo=empresa_periodo,
        estatus__in=[1, True],
        fecha__range=[fecha_desde, fecha_hasta]
    ).order_by('fecha', 'correlativo')

    return empresa, periodo, fecha_desde, fecha_hasta, asientos, empresa_periodo, None


@login_required
def libro_diario_excel(request):
    empresa, periodo, fecha_desde, fecha_hasta, asientos, empresa_periodo, _ = _get_libro_diario_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    if not asientos.exists():
        messages.warning(request, 'No hay datos para exportar')
        return redirect('libro_diario')

    datos  = LibroDiarioService.get_datos_reporte(asientos, fecha_desde, fecha_hasta, empresa.razon_social, periodo.nombre)
    paginas = LibroDiarioService.paginar_con_van_vienen(datos)
    total_debe_general  = sum(b['total_debe']  for b in datos if b['tipo'] == 'asiento')
    total_haber_general = sum(b['total_haber'] for b in datos if b['tipo'] == 'asiento')

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Diario"

    # Estilos
    fnt_titulo   = Font(bold=True, size=13)
    fnt_bold     = Font(bold=True, size=10)
    fnt_normal   = Font(size=10)
    fnt_italica  = Font(size=10, italic=True)
    fnt_detalle  = Font(size=9, italic=True, color='444444')

    aln_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    aln_right  = Alignment(horizontal='right',  vertical='center')
    aln_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    borde_grueso_top = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='medium'), bottom=Side(style='thin')
    )
    borde_grueso_bot = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='medium')
    )

    fill_header   = PatternFill(start_color='1B2B4E', end_color='1B2B4E', fill_type='solid')
    fill_mes      = PatternFill(start_color='243B67', end_color='243B67', fill_type='solid')
    fill_fecha    = PatternFill(start_color='EEF2F7', end_color='EEF2F7', fill_type='solid')
    fill_comentario = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    fill_van      = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
    fill_total    = PatternFill(start_color='1B2B4E', end_color='1B2B4E', fill_type='solid')
    fill_detalle  = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    fnt_header_white = Font(bold=True, size=10, color='FFFFFF')
    fnt_mes_white    = Font(bold=True, size=10, color='FFFFFF')
    fnt_total_white  = Font(bold=True, size=10, color='FFFFFF')

    # Columnas: A=Partida(10), B=Cuenta(50), C=Debe(16), D=Haber(16)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    fila = 1

    def celda(r, c, valor='', fuente=None, alineacion=None, relleno=None, bordes=None, formato=None):
        cl = ws.cell(row=r, column=c, value=valor)
        if fuente:     cl.font      = fuente
        if alineacion: cl.alignment = alineacion
        if relleno:    cl.fill      = relleno
        if bordes:     cl.border    = bordes
        if formato:    cl.number_format = formato
        return cl

    def fila_merge(r, texto, fuente, relleno, altura=14):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        cl = ws.cell(row=r, column=1, value=texto)
        cl.font = fuente
        cl.alignment = aln_center
        cl.fill = relleno
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = borde
        ws.row_dimensions[r].height = altura

    for pagina in paginas:
        # Encabezado de página
        fila_merge(fila, 'LIBRO DIARIO', fnt_titulo, PatternFill(fill_type=None))
        fila += 1
        nombre_emp = empresa.nombre_comercial or empresa.razon_social
        fila_merge(fila, nombre_emp, fnt_bold, PatternFill(fill_type=None))
        fila += 1
        fila_merge(fila,
            f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')} — "
            f"Período: {periodo.nombre} — (Expresado en Quetzales) — "
            f"Página {pagina['numero']} de {len(paginas)}",
            Font(size=9), PatternFill(fill_type=None), altura=12)
        fila += 1

        # Encabezados de columna
        for c, txt in enumerate(['PARTIDA', 'NOMBRE DE LA CUENTA', 'DEBE', 'HABER'], 1):
            celda(fila, c, txt, fuente=fnt_header_white, alineacion=aln_center, relleno=fill_header, bordes=borde)
        ws.row_dimensions[fila].height = 15
        fila += 1

        # VIENEN
        if pagina['corte_partida_anterior']:
            celda(fila, 1, '', bordes=borde, relleno=fill_van)
            celda(fila, 2, 'VIENEN', fuente=fnt_bold, alineacion=aln_right, relleno=fill_van, bordes=borde)
            celda(fila, 3, float(pagina['vienen_debe']),  fuente=fnt_bold, alineacion=aln_right, relleno=fill_van, bordes=borde, formato='#,##0.00')
            celda(fila, 4, float(pagina['vienen_haber']), fuente=fnt_bold, alineacion=aln_right, relleno=fill_van, bordes=borde, formato='#,##0.00')
            fila += 1

        # Registros
        for reg in pagina['registros']:
            tipo = reg['tipo']

            if tipo == 'separador_mes':
                fila_merge(fila, reg['texto'], fnt_mes_white, fill_mes, altura=13)
                fila += 1

            elif tipo == 'fecha':
                fila_merge(fila, reg['texto'], fnt_bold, fill_fecha, altura=13)
                fila += 1

            elif tipo == 'espacio':
                for c in range(1, 5):
                    ws.cell(row=fila, column=c).border = Border()
                ws.row_dimensions[fila].height = 5
                fila += 1

            elif tipo == 'movimiento':
                partida_txt = f"---{reg['correlativo']}---" if reg['correlativo'] else ''
                celda(fila, 1, partida_txt, fuente=fnt_bold, alineacion=aln_center, bordes=borde)
                celda(fila, 2, reg['cuenta_nombre'], fuente=fnt_normal, alineacion=aln_left, bordes=borde)
                celda(fila, 3, float(reg['debe'])  if reg['debe']  > 0 else '', fuente=fnt_normal, alineacion=aln_right, bordes=borde, formato='#,##0.00')
                celda(fila, 4, float(reg['haber']) if reg['haber'] > 0 else '', fuente=fnt_normal, alineacion=aln_right, bordes=borde, formato='#,##0.00')
                fila += 1

            elif tipo == 'detalle':
                celda(fila, 1, '', bordes=borde, relleno=fill_detalle)
                celda(fila, 2, f'  {reg["cuenta_nombre"]}', fuente=fnt_detalle, alineacion=aln_left, relleno=fill_detalle, bordes=borde)
                celda(fila, 3, float(reg['debe'])  if reg['debe']  > 0 else '', fuente=fnt_detalle, alineacion=aln_right, relleno=fill_detalle, bordes=borde, formato='#,##0.00')
                celda(fila, 4, float(reg['haber']) if reg['haber'] > 0 else '', fuente=fnt_detalle, alineacion=aln_right, relleno=fill_detalle, bordes=borde, formato='#,##0.00')
                fila += 1

            elif tipo == 'comentario':
                celda(fila, 1, '', relleno=fill_comentario, bordes=borde_grueso_top)
                celda(fila, 2, f"v/ {reg['cuenta_nombre']}", fuente=fnt_italica, alineacion=aln_left, relleno=fill_comentario, bordes=borde_grueso_top)
                celda(fila, 3, float(reg['debe']),  fuente=fnt_bold, alineacion=aln_right, relleno=fill_comentario, bordes=borde_grueso_top, formato='#,##0.00')
                celda(fila, 4, float(reg['haber']), fuente=fnt_bold, alineacion=aln_right, relleno=fill_comentario, bordes=borde_grueso_top, formato='#,##0.00')
                fila += 1

        # VAN
        if not pagina['es_ultima'] and pagina['corte_partida']:
            celda(fila, 1, '', bordes=borde, relleno=fill_van)
            celda(fila, 2, 'VAN', fuente=fnt_bold, alineacion=aln_right, relleno=fill_van, bordes=borde)
            celda(fila, 3, float(pagina['van_debe']),  fuente=fnt_bold, alineacion=aln_right, relleno=fill_van, bordes=borde, formato='#,##0.00')
            celda(fila, 4, float(pagina['van_haber']), fuente=fnt_bold, alineacion=aln_right, relleno=fill_van, bordes=borde, formato='#,##0.00')
            fila += 1
            # Salto de página Excel
            from openpyxl.worksheet.pagebreak import Break
            ws.row_breaks.append(Break(id=fila))
            fila += 3

    # Totales generales
    fila += 1
    celda(fila, 1, '', relleno=fill_total, bordes=borde)
    celda(fila, 2, 'TOTALES GENERALES DEL PERÍODO', fuente=fnt_total_white, alineacion=aln_right, relleno=fill_total, bordes=borde)
    celda(fila, 3, float(total_debe_general),  fuente=fnt_total_white, alineacion=aln_right, relleno=fill_total, bordes=borde, formato='#,##0.00')
    celda(fila, 4, float(total_haber_general), fuente=fnt_total_white, alineacion=aln_right, relleno=fill_total, bordes=borde, formato='#,##0.00')

    # Configuración de impresión
    ws.page_setup.paperSize  = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    nombre = f"LibroDiario_{(empresa.nombre_comercial or empresa.razon_social).replace(' ','_')}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response


@login_required
def libro_diario_pdf(request):
    from io import BytesIO
    try:
        from xhtml2pdf import pisa
    except ImportError:
        messages.error(request, 'xhtml2pdf no está instalado. Ejecute: pip install xhtml2pdf')
        return redirect('libro_diario')

    empresa, periodo, fecha_desde, fecha_hasta, asientos, empresa_periodo, _ = _get_libro_diario_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    if not asientos.exists():
        messages.warning(request, 'No hay datos para exportar')
        return redirect('libro_diario')

    datos   = LibroDiarioService.get_datos_reporte(asientos, fecha_desde, fecha_hasta, empresa.razon_social, periodo.nombre)
    paginas = LibroDiarioService.paginar_con_van_vienen(datos)
    total_debe_general  = sum(b['total_debe']  for b in datos if b['tipo'] == 'asiento')
    total_haber_general = sum(b['total_haber'] for b in datos if b['tipo'] == 'asiento')

    nombre_empresa = empresa.razon_social or empresa.nombre_comercial

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  @page {{
    size: letter portrait;
    margin: 1.5cm 1.8cm;
  }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 9pt; color: #000; }}
  .page-break {{ page-break-before: always; }}
  .enc-tabla {{ width: 100%; border-collapse: collapse; margin-bottom: 4pt; border: none; }}
  .enc-izq {{ text-align: center; vertical-align: middle; width: 85%; padding: 0; }}
  .enc-der {{ text-align: right; vertical-align: top; padding: 0; font-size: 8.5pt; color: #333; white-space: nowrap; width: 15%; }}
  .enc-titulo  {{ font-size: 13pt; font-weight: bold; margin: 0; text-align: center; }}
  .enc-empresa {{ font-size: 10pt; font-weight: bold; margin: 1pt 0 0 0; text-align: center; }}
  .enc-fechas  {{ font-size: 8.5pt; color: #333; margin: 1pt 0 0 0; }}
  .enc-moneda  {{ font-size: 8pt; color: #555; margin: 1pt 0 0 0; }}
  table {{ border-collapse: collapse; margin-top: 3px; width: 97%; margin-left: 1.5%; margin-right: 1.5%; }}
  th {{ background-color: #EEF2F7; color: #0A1628; font-size: 9pt; padding: 2px 5px;
        border-left: 0.5pt solid #94a3b8; border-right: 0.5pt solid #94a3b8;
        border-top: 0.5pt solid #94a3b8; border-bottom: 1pt solid #94a3b8;
        text-align: center; font-weight: bold; }}
  td {{ border-left: 0.5pt solid #94a3b8; border-right: 0.5pt solid #94a3b8;
        border-top: none; border-bottom: none;
        padding: 1px 4px; vertical-align: middle; font-size: 8.5pt; }}
  .enc-tabla td {{ border: none; padding: 0 2pt; background: white; }}
  .col-partida  {{ width: 40pt; text-align: center; }}
  .col-cuenta   {{ text-align: left; }}
  .col-auxiliar {{ width: 55pt; text-align: right; }}
  .col-monto    {{ width: 60pt; text-align: right; }}
  .row-mes td  {{ background-color: #D1D5DB; color: #0A1628; font-weight: bold;
                  text-align: center; font-size: 9pt;
                  border: 0.5pt solid #94a3b8; padding: 1px 5px; line-height: 1.2; }}
  .row-fecha td {{ background-color: #D1D5DB; font-weight: bold; text-align: center;
                   padding: 1px 4px; border: 0.5pt solid #94a3b8; line-height: 1.2; }}
  .row-detalle td {{ font-style: italic; color: #444; padding-left: 14pt; }}
  .row-comentario td {{ background-color: #EEF2F7; font-style: italic;
                        border-top: 0.8pt solid #000; border-bottom: 1pt solid #000; }}
  .row-comentario .col-monto, .row-comentario .col-auxiliar {{ font-weight: bold; font-style: normal; }}
  .row-espacio td {{ border: none; height: 2pt; background: white; }}
  .row-van td, .row-vienen td {{ background-color: #EEF2F7; font-weight: bold;
                                  border-top: 1pt solid #000; border-bottom: 1pt solid #000; }}
  .row-van .col-cuenta, .row-vienen .col-cuenta {{ text-align: right; }}
  .row-total td {{ background-color: #1B2B4E; color: white; font-weight: bold;
                   border-top: 1pt solid #000; border-bottom: 1pt solid #000; }}
  .row-total .col-cuenta {{ text-align: right; }}
</style>
</head>
<body>
"""

    for i, pagina in enumerate(paginas):
        if i > 0:
            html += '<div class="page-break"></div>'

        html += f"""
<table class="enc-tabla">
  <tr>
    <td class="enc-izq">
      <div class="enc-titulo">LIBRO DE DIARIO</div>
      <div class="enc-empresa">{nombre_empresa}</div>
      <div class="enc-fechas">Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}</div>
      <div class="enc-moneda">(Expresado en Quetzales)</div>
    </td>
    <td class="enc-der">Folio No. {pagina['numero']}</td>
  </tr>
</table>
<table>
<thead>
  <tr>
    <th class="col-partida">PARTIDA</th>
    <th class="col-cuenta">NOMBRE DE LA CUENTA</th>
    <th class="col-auxiliar"></th>
    <th class="col-monto">DEBE</th>
    <th class="col-monto">HABER</th>
  </tr>
</thead>
<tbody>
"""
        if not pagina['es_primera'] and paginas[i-1]['corte_partida']:
            html += f"""<tr class="row-vienen">
  <td class="col-partida"></td>
  <td class="col-cuenta" style="text-align:right;">VIENEN</td>
  <td class="col-auxiliar"></td>
  <td class="col-monto">{pagina['vienen_debe']:,.2f}</td>
  <td class="col-monto">{pagina['vienen_haber']:,.2f}</td>
</tr>"""

        for reg in pagina['registros']:
            tipo = reg['tipo']

            if tipo == 'separador_mes':
                html += f'<tr class="row-mes"><td colspan="5">{reg["texto"]}</td></tr>'

            elif tipo == 'fecha':
                html += f'<tr class="row-fecha"><td colspan="5">{reg["texto"]}</td></tr>'

            elif tipo == 'espacio':
                html += '<tr class="row-espacio"><td colspan="5"></td></tr>'

            elif tipo == 'movimiento':
                partida = str(reg['correlativo']) if reg['correlativo'] else ''
                debe  = f"{reg['debe']:,.2f}"  if reg['debe']  > 0 else ''
                haber = f"{reg['haber']:,.2f}" if reg['haber'] > 0 else ''
                html += f"""<tr>
  <td class="col-partida">{partida}</td>
  <td class="col-cuenta">{reg['cuenta_nombre']}</td>
  <td class="col-auxiliar"></td>
  <td class="col-monto">{debe}</td>
  <td class="col-monto">{haber}</td>
</tr>"""

            elif tipo == 'detalle':
                aux = f"{reg['debe']:,.2f}" if reg['debe'] > 0 else f"{reg['haber']:,.2f}" if reg['haber'] > 0 else ''
                html += f"""<tr class="row-detalle">
  <td class="col-partida"></td>
  <td class="col-cuenta">{reg['cuenta_nombre']}</td>
  <td class="col-auxiliar">{aux}</td>
  <td class="col-monto"></td>
  <td class="col-monto"></td>
</tr>"""

            elif tipo == 'comentario':
                html += f"""<tr class="row-comentario">
  <td class="col-partida"></td>
  <td class="col-cuenta">v/ {reg['cuenta_nombre']}</td>
  <td class="col-auxiliar"></td>
  <td class="col-monto">{reg['debe']:,.2f}</td>
  <td class="col-monto">{reg['haber']:,.2f}</td>
</tr>"""

        if not pagina['es_ultima'] and pagina['corte_partida']:
            html += f"""<tr class="row-van">
  <td class="col-partida"></td>
  <td class="col-cuenta" style="text-align:right;">VAN</td>
  <td class="col-auxiliar"></td>
  <td class="col-monto">{pagina['van_debe']:,.2f}</td>
  <td class="col-monto">{pagina['van_haber']:,.2f}</td>
</tr>"""

        html += '</tbody></table>'

    html += f"""
<table style="margin-top:6pt;">
<tbody>
<tr class="row-total">
  <td class="col-partida"></td>
  <td class="col-cuenta">TOTALES GENERALES DEL PERIODO</td>
  <td class="col-auxiliar"></td>
  <td class="col-monto">{total_debe_general:,.2f}</td>
  <td class="col-monto">{total_haber_general:,.2f}</td>
</tr>
</tbody>
</table>
</body></html>"""

    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=buffer, encoding='utf-8')

    if pisa_status.err:
        messages.error(request, 'Error al generar el PDF')
        return redirect('libro_diario')

    buffer.seek(0)
    nombre = f"LibroDiario_{nombre_empresa.replace(' ','_')}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response





# ==================== ESTADO DE RESULTADOS ====================

def _get_estado_resultados_datos(request):
    from administracion.services.estado_resultados_service import EstadoResultadosService

    empresa_activa_id = request.session.get('empresa_activa_id')
    empresa_periodo_activo_id = request.session.get('empresa_periodo_activo_id')

    if not empresa_activa_id or not empresa_periodo_activo_id:
        return None, None, None, None, None

    try:
        empresa = Empresa.objects.get(id=empresa_activa_id)
        empresa_periodo = EmpresaPeriodo.objects.select_related('id_periodo').get(id=empresa_periodo_activo_id)
        periodo = empresa_periodo.id_periodo
    except (Empresa.DoesNotExist, EmpresaPeriodo.DoesNotExist):
        return None, None, None, None, None

    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    try:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date() if fecha_desde else periodo.fecha_inicial
    except ValueError:
        fecha_desde = periodo.fecha_inicial

    try:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date() if fecha_hasta else periodo.fecha_final
    except ValueError:
        fecha_hasta = periodo.fecha_final

    datos = EstadoResultadosService.get_datos_reporte(empresa_periodo, fecha_desde, fecha_hasta)
    return empresa, periodo, fecha_desde, fecha_hasta, datos


@login_required
def estado_resultados(request):
    empresa, periodo, fecha_desde, fecha_hasta, datos = _get_estado_resultados_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    sin_datos = not datos['ingresos']['grupos'] and not datos['egresos']['grupos']

    return render(request, 'administracion/reportes/estado_resultados.html', {
        'sin_datos':     sin_datos,
        'empresa':       empresa,
        'periodo':       periodo,
        'fecha_desde':   fecha_desde,
        'fecha_hasta':   fecha_hasta,
        'datos':         datos,
        'fecha_reporte': timezone.now(),
    })


@login_required
def estado_resultados_excel(request):
    from decimal import Decimal

    empresa, periodo, fecha_desde, fecha_hasta, datos = _get_estado_resultados_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    if not datos['ingresos']['grupos'] and not datos['egresos']['grupos']:
        messages.warning(request, 'No hay datos para exportar')
        return redirect('estado_resultados')

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    fnt_titulo = Font(bold=True, size=13)
    fnt_bold   = Font(bold=True, size=10)
    fnt_normal = Font(size=10)

    aln_center = Alignment(horizontal='center', vertical='center')
    aln_right  = Alignment(horizontal='right',  vertical='center')
    aln_left   = Alignment(horizontal='left',   vertical='center')

    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    fill_header  = PatternFill(start_color='1B2B4E', end_color='1B2B4E', fill_type='solid')
    fill_sg      = PatternFill(start_color='EEF2F7', end_color='EEF2F7', fill_type='solid')
    fill_sg_tot  = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    fill_sec_tot = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
    fill_util    = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    fill_perd    = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    fill_alt     = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    fnt_white = Font(bold=True, size=10, color='FFFFFF')
    fnt_util  = Font(bold=True, size=11, color='065F46')
    fnt_perd  = Font(bold=True, size=11, color='991B1B')

    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 18

    nombre_empresa = empresa.nombre_comercial or empresa.razon_social
    fila = 1

    def merge(r, val, fuente, relleno, altura=14):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cl = ws.cell(row=r, column=1, value=val)
        cl.font = fuente; cl.alignment = aln_center; cl.fill = relleno
        for c in range(1, 3): ws.cell(row=r, column=c).border = borde
        ws.row_dimensions[r].height = altura

    def cel(r, c, val='', fuente=None, aln=None, relleno=None, fmt=None):
        cl = ws.cell(row=r, column=c, value=val)
        if fuente:  cl.font      = fuente
        if aln:     cl.alignment = aln
        if relleno: cl.fill      = relleno
        if fmt:     cl.number_format = fmt
        cl.border = borde
        return cl

    merge(fila, 'ESTADO DE RESULTADOS', fnt_titulo, PatternFill(fill_type=None), 18); fila += 1
    merge(fila, nombre_empresa, fnt_bold, PatternFill(fill_type=None)); fila += 1
    merge(fila,
        f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')} — (Expresado en Quetzales)",
        Font(size=9), PatternFill(fill_type=None), 12)
    fila += 2

    def escribir_seccion(titulo, grupos, total):
        nonlocal fila
        merge(fila, titulo, fnt_white, fill_header, 14); fila += 1

        for grupo in grupos:
            if len(grupo['cuentas']) > 1:
                merge(fila, grupo['nombre'], fnt_bold, fill_sg, 13); fila += 1

            for i, c in enumerate(grupo['cuentas']):
                rf = fill_alt if i % 2 == 1 else None
                cel(fila, 1, c['cuenta'].nombre, fuente=fnt_normal, aln=aln_left, relleno=rf)
                cel(fila, 2, float(c['monto']), fuente=fnt_normal, aln=aln_right, relleno=rf, fmt='#,##0.00')
                fila += 1

            if len(grupo['cuentas']) > 1:
                cel(fila, 1, f"Subtotal {grupo['nombre']}", fuente=fnt_bold, aln=aln_right, relleno=fill_sg_tot)
                cel(fila, 2, float(grupo['subtotal']), fuente=fnt_bold, aln=aln_right, relleno=fill_sg_tot, fmt='#,##0.00')
                fila += 1

        cel(fila, 1, f"TOTAL {titulo}", fuente=fnt_bold, aln=aln_right, relleno=fill_sec_tot)
        cel(fila, 2, float(total), fuente=fnt_bold, aln=aln_right, relleno=fill_sec_tot, fmt='#,##0.00')
        fila += 2

    escribir_seccion('INGRESOS', datos['ingresos']['grupos'], datos['ingresos']['total'])
    escribir_seccion('EGRESOS',  datos['egresos']['grupos'],  datos['egresos']['total'])

    lbl = 'UTILIDAD DEL EJERCICIO' if datos['es_utilidad'] else 'PÉRDIDA DEL EJERCICIO'
    fill_r = fill_util if datos['es_utilidad'] else fill_perd
    fnt_r  = fnt_util  if datos['es_utilidad'] else fnt_perd
    val_r  = float(datos['utilidad']) * (1 if datos['es_utilidad'] else -1)
    cel(fila, 1, lbl, fuente=fnt_r, aln=aln_right, relleno=fill_r)
    cel(fila, 2, val_r, fuente=fnt_r, aln=aln_right, relleno=fill_r, fmt='#,##0.00')

    ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    nombre = f"EstadoResultados_{nombre_empresa.replace(' ','_')}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response


@login_required
def estado_resultados_pdf(request):
    from io import BytesIO
    from decimal import Decimal
    try:
        from xhtml2pdf import pisa
    except ImportError:
        messages.error(request, 'xhtml2pdf no está instalado.')
        return redirect('estado_resultados')

    empresa, periodo, fecha_desde, fecha_hasta, datos = _get_estado_resultados_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    nombre_empresa = empresa.razon_social or empresa.nombre_comercial

    def render_seccion(titulo, grupos, total):
        html = f'<tr class="sec-hdr"><td colspan="2">{titulo}</td></tr>'
        for grupo in grupos:
            if len(grupo['cuentas']) > 1:
                html += f'<tr class="sg-hdr"><td colspan="2">{grupo["nombre"]}</td></tr>'
            for j, c in enumerate(grupo['cuentas']):
                cls = 'alt' if j % 2 else ''
                html += f'<tr class="{cls}"><td class="ind">{c["cuenta"].nombre}</td><td class="num">{c["monto"]:,.2f}</td></tr>'
            if len(grupo['cuentas']) > 1:
                html += f'<tr class="sg-tot"><td style="text-align:right;">Subtotal {grupo["nombre"]}</td><td class="num">{grupo["subtotal"]:,.2f}</td></tr>'
        html += f'<tr class="sec-tot"><td style="text-align:right;">TOTAL {titulo}</td><td class="num">{total:,.2f}</td></tr>'
        html += '<tr class="sep"><td colspan="2"></td></tr>'
        return html

    lbl_res = 'UTILIDAD DEL EJERCICIO' if datos['es_utilidad'] else 'PERDIDA DEL EJERCICIO'
    cls_res = 'util' if datos['es_utilidad'] else 'perd'
    val_res = datos['utilidad']
    fmt_res = f"{val_res:,.2f}" if datos['es_utilidad'] else f"({val_res:,.2f})"

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: letter portrait; margin: 1.5cm 1.8cm; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 8.5pt; color: #000; }}
  .enc-tabla {{ width: 100%; border-collapse: collapse; margin-bottom: 6pt; border: none; }}
  .enc-tabla td {{ border: none; padding: 0 2pt; background: white; }}
  .enc-titulo  {{ font-size: 13pt; font-weight: bold; text-align: center; margin: 0; }}
  .enc-empresa {{ font-size: 10pt; font-weight: bold; text-align: center; margin: 1pt 0 0 0; }}
  .enc-fechas  {{ font-size: 8.5pt; color: #333; text-align: center; margin: 1pt 0 0 0; }}
  .enc-moneda  {{ font-size: 8pt; color: #555; text-align: center; margin: 1pt 0 0 0; }}
  table.detalle {{ width: 70%; margin: 0 auto; border-collapse: collapse; }}
  td {{ border-left: 0.5pt solid #94a3b8; border-right: 0.5pt solid #94a3b8;
        border-top: none; border-bottom: none;
        padding: 2px 6px; font-size: 8pt; }}
  .enc-tabla td {{ border: none; padding: 0 2pt; background: white; }}
  .num {{ text-align: right; width: 80pt; }}
  .ind {{ padding-left: 16pt; }}
  .sec-hdr td {{ background-color: #D1D5DB; color: #0A1628; font-weight: bold;
                  font-size: 8.5pt; padding: 3px 6px;
                  border: 0.5pt solid #94a3b8; border-bottom: 1pt solid #94a3b8; }}
  .sg-hdr  td {{ background-color: #EEF2F7; font-weight: bold; padding: 2px 6px;
                  border: 0.5pt solid #94a3b8; }}
  .alt     td {{ background-color: #FAFAFA; }}
  .sg-tot  td {{ background-color: #EEF2F7; font-weight: bold;
                  border-top: 0.8pt solid #000; border-bottom: none; }}
  .sec-tot td {{ background-color: #D1D5DB; font-weight: bold;
                  border-top: 1pt solid #000; border-bottom: 1pt solid #000; font-size: 8.5pt; }}
  .sep     td {{ border: none; height: 6pt; background: white; }}
  .util    td {{ background-color: #D1FAE5; color: #065F46; font-weight: bold;
                  border-top: 1.5pt solid #000; border-bottom: 2pt double #000; font-size: 9pt; }}
  .perd    td {{ background-color: #FEE2E2; color: #991B1B; font-weight: bold;
                  border-top: 1.5pt solid #000; border-bottom: 2pt double #000; font-size: 9pt; }}
</style></head><body>
<table class="enc-tabla">
  <tr>
    <td>
      <div class="enc-titulo">ESTADO DE RESULTADOS</div>
      <div class="enc-empresa">{nombre_empresa}</div>
      <div class="enc-fechas">Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}</div>
      <div class="enc-moneda">(Expresado en Quetzales)</div>
    </td>
  </tr>
</table>
<table class="detalle">
<tbody>
{render_seccion('INGRESOS', datos['ingresos']['grupos'], datos['ingresos']['total'])}
{render_seccion('EGRESOS',  datos['egresos']['grupos'],  datos['egresos']['total'])}
<tr class="{cls_res}">
  <td style="text-align:right;">{lbl_res}</td>
  <td class="num">{fmt_res}</td>
</tr>
</tbody>
</table>
</body></html>"""

    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
    if pisa_status.err:
        messages.error(request, 'Error al generar el PDF')
        return redirect('estado_resultados')

    buffer.seek(0)
    nombre = f"EstadoResultados_{nombre_empresa.replace(' ','_')}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response


# ==================== BALANCE GENERAL ====================

def _get_balance_general_datos(request):
    from administracion.services.balance_general_service import BalanceGeneralService

    empresa_activa_id = request.session.get('empresa_activa_id')
    empresa_periodo_activo_id = request.session.get('empresa_periodo_activo_id')

    if not empresa_activa_id or not empresa_periodo_activo_id:
        return None, None, None, None

    try:
        empresa = Empresa.objects.get(id=empresa_activa_id)
        empresa_periodo = EmpresaPeriodo.objects.select_related('id_periodo').get(id=empresa_periodo_activo_id)
        periodo = empresa_periodo.id_periodo
    except (Empresa.DoesNotExist, EmpresaPeriodo.DoesNotExist):
        return None, None, None, None

    fecha_hasta = request.GET.get('fecha_hasta')
    try:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date() if fecha_hasta else periodo.fecha_final
    except ValueError:
        fecha_hasta = periodo.fecha_final

    datos = BalanceGeneralService.get_datos_reporte(empresa_periodo, fecha_hasta)
    return empresa, periodo, fecha_hasta, datos


@login_required
def balance_general(request):
    empresa, periodo, fecha_hasta, datos = _get_balance_general_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    sin_datos = (
        not datos['activo']['grupos'] and
        not datos['pasivo']['grupos'] and
        not datos['capital']['grupos'] and
        datos['capital']['utilidad_ejercicio'] == 0
    )

    from decimal import Decimal
    diferencia = abs(datos['activo']['total'] - datos['total_pasivo_capital'])

    return render(request, 'administracion/reportes/balance_general.html', {
        'sin_datos':    sin_datos,
        'empresa':      empresa,
        'periodo':      periodo,
        'fecha_hasta':  fecha_hasta,
        'datos':        datos,
        'diferencia':   diferencia,
        'fecha_reporte': timezone.now(),
    })


@login_required
def balance_general_excel(request):
    from decimal import Decimal

    empresa, periodo, fecha_hasta, datos = _get_balance_general_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance General"

    fnt_titulo = Font(bold=True, size=13)
    fnt_bold   = Font(bold=True, size=10)
    fnt_normal = Font(size=10)

    aln_center = Alignment(horizontal='center', vertical='center')
    aln_right  = Alignment(horizontal='right',  vertical='center')
    aln_left   = Alignment(horizontal='left',   vertical='center')

    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    fill_header  = PatternFill(start_color='1B2B4E', end_color='1B2B4E', fill_type='solid')
    fill_sec     = PatternFill(start_color='EEF2F7', end_color='EEF2F7', fill_type='solid')
    fill_subtot  = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    fill_total   = PatternFill(start_color='1B2B4E', end_color='1B2B4E', fill_type='solid')
    fill_subtot2 = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
    fill_alt     = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    fill_util    = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    fill_perd    = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    fnt_white     = Font(bold=True, size=10, color='FFFFFF')
    fnt_util      = Font(bold=True, size=10, color='065F46')
    fnt_perd      = Font(bold=True, size=10, color='991B1B')

    # Columnas: A-B = Activo, C vacía, D-E = Pasivo+Capital
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 38
    ws.column_dimensions['E'].width = 16

    nombre_empresa = empresa.nombre_comercial or empresa.razon_social
    fila = 1

    def cel(r, c, val='', fuente=None, aln=None, relleno=None, fmt=None, merge_end=None):
        if merge_end:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_end)
        cl = ws.cell(row=r, column=c, value=val)
        if fuente:  cl.font      = fuente
        if aln:     cl.alignment = aln
        if relleno: cl.fill      = relleno
        if fmt:     cl.number_format = fmt
        cl.border = borde
        return cl

    def encabezado_col(r, col_ini, titulo):
        ws.merge_cells(start_row=r, start_column=col_ini, end_row=r, end_column=col_ini+1)
        cl = ws.cell(row=r, column=col_ini, value=titulo)
        cl.font = fnt_white; cl.alignment = aln_center; cl.fill = fill_header
        ws.cell(row=r, column=col_ini).border = borde
        ws.cell(row=r, column=col_ini+1).border = borde

    # Título
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    cl = ws.cell(row=fila, column=1, value='BALANCE GENERAL')
    cl.font = fnt_titulo; cl.alignment = aln_center
    for c in range(1, 6): ws.cell(row=fila, column=c).border = borde
    ws.row_dimensions[fila].height = 18; fila += 1

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    cl = ws.cell(row=fila, column=1, value=nombre_empresa)
    cl.font = fnt_bold; cl.alignment = aln_center
    for c in range(1, 6): ws.cell(row=fila, column=c).border = borde; fila += 1

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    cl = ws.cell(row=fila, column=1, value=f"Al {fecha_hasta.strftime('%d/%m/%Y')} — (Expresado en Quetzales)")
    cl.font = Font(size=9); cl.alignment = aln_center
    for c in range(1, 6): ws.cell(row=fila, column=c).border = borde
    fila += 2

    # Encabezados de columna
    encabezado_col(fila, 1, 'ACTIVO')
    ws.cell(row=fila, column=3).border = borde
    encabezado_col(fila, 4, 'PASIVO + CAPITAL')
    ws.row_dimensions[fila].height = 14; fila += 1

    fila_inicio = fila

    # ── ACTIVO ──
    fila_act = fila
    for grupo in datos['activo']['grupos']:
        ws.merge_cells(start_row=fila_act, start_column=1, end_row=fila_act, end_column=2)
        cl = ws.cell(row=fila_act, column=1, value=grupo['nombre'].upper())
        cl.font = fnt_bold; cl.fill = fill_sec; cl.border = borde
        ws.cell(row=fila_act, column=2).fill = fill_sec; ws.cell(row=fila_act, column=2).border = borde
        fila_act += 1

        for i, c in enumerate(grupo['cuentas']):
            rf = fill_alt if i % 2 == 1 else None
            cel(fila_act, 1, c['cuenta'].nombre, fuente=fnt_normal, aln=aln_left, relleno=rf)
            cel(fila_act, 2, float(c['saldo']), fuente=fnt_normal, aln=aln_right, relleno=rf, fmt='#,##0.00')
            fila_act += 1

        cel(fila_act, 1, f"Total {grupo['nombre']}", fuente=fnt_bold, aln=aln_right, relleno=fill_subtot)
        cel(fila_act, 2, float(grupo['subtotal']), fuente=fnt_bold, aln=aln_right, relleno=fill_subtot, fmt='#,##0.00')
        fila_act += 1

    cel(fila_act, 1, 'TOTAL ACTIVO', fuente=fnt_white, aln=aln_right, relleno=fill_total)
    cel(fila_act, 2, float(datos['activo']['total']), fuente=fnt_white, aln=aln_right, relleno=fill_total, fmt='#,##0.00')
    fila_act += 1

    # ── PASIVO + CAPITAL ──
    fila_pas = fila

    # Encabezado PASIVO
    ws.merge_cells(start_row=fila_pas, start_column=4, end_row=fila_pas, end_column=5)
    cl = ws.cell(row=fila_pas, column=4, value='PASIVO')
    cl.font = fnt_white; cl.fill = fill_header; cl.alignment = aln_center; cl.border = borde
    ws.cell(row=fila_pas, column=5).fill = fill_header; ws.cell(row=fila_pas, column=5).border = borde
    fila_pas += 1

    for grupo in datos['pasivo']['grupos']:
        ws.merge_cells(start_row=fila_pas, start_column=4, end_row=fila_pas, end_column=5)
        cl = ws.cell(row=fila_pas, column=4, value=grupo['nombre'].upper())
        cl.font = fnt_bold; cl.fill = fill_sec; cl.border = borde
        ws.cell(row=fila_pas, column=5).fill = fill_sec; ws.cell(row=fila_pas, column=5).border = borde
        fila_pas += 1

        for i, c in enumerate(grupo['cuentas']):
            rf = fill_alt if i % 2 == 1 else None
            cel(fila_pas, 4, c['cuenta'].nombre, fuente=fnt_normal, aln=aln_left, relleno=rf)
            cel(fila_pas, 5, float(c['saldo']), fuente=fnt_normal, aln=aln_right, relleno=rf, fmt='#,##0.00')
            fila_pas += 1

        cel(fila_pas, 4, f"Total {grupo['nombre']}", fuente=fnt_bold, aln=aln_right, relleno=fill_subtot)
        cel(fila_pas, 5, float(grupo['subtotal']), fuente=fnt_bold, aln=aln_right, relleno=fill_subtot, fmt='#,##0.00')
        fila_pas += 1

    cel(fila_pas, 4, 'TOTAL PASIVO', fuente=fnt_bold, aln=aln_right, relleno=fill_subtot2)
    cel(fila_pas, 5, float(datos['pasivo']['total']), fuente=fnt_bold, aln=aln_right, relleno=fill_subtot2, fmt='#,##0.00')
    fila_pas += 1

    # Encabezado CAPITAL
    ws.merge_cells(start_row=fila_pas, start_column=4, end_row=fila_pas, end_column=5)
    cl = ws.cell(row=fila_pas, column=4, value='CAPITAL')
    cl.font = fnt_white; cl.fill = fill_header; cl.alignment = aln_center; cl.border = borde
    ws.cell(row=fila_pas, column=5).fill = fill_header; ws.cell(row=fila_pas, column=5).border = borde
    fila_pas += 1

    for grupo in datos['capital']['grupos']:
        ws.merge_cells(start_row=fila_pas, start_column=4, end_row=fila_pas, end_column=5)
        cl = ws.cell(row=fila_pas, column=4, value=grupo['nombre'].upper())
        cl.font = fnt_bold; cl.fill = fill_sec; cl.border = borde
        ws.cell(row=fila_pas, column=5).fill = fill_sec; ws.cell(row=fila_pas, column=5).border = borde
        fila_pas += 1

        for i, c in enumerate(grupo['cuentas']):
            rf = fill_alt if i % 2 == 1 else None
            cel(fila_pas, 4, c['cuenta'].nombre, fuente=fnt_normal, aln=aln_left, relleno=rf)
            cel(fila_pas, 5, float(c['saldo']), fuente=fnt_normal, aln=aln_right, relleno=rf, fmt='#,##0.00')
            fila_pas += 1

        cel(fila_pas, 4, f"Total {grupo['nombre']}", fuente=fnt_bold, aln=aln_right, relleno=fill_subtot)
        cel(fila_pas, 5, float(grupo['subtotal']), fuente=fnt_bold, aln=aln_right, relleno=fill_subtot, fmt='#,##0.00')
        fila_pas += 1

    # Utilidad/Pérdida
    lbl_util = 'Utilidad del Ejercicio' if datos['capital']['es_utilidad'] else 'Pérdida del Ejercicio'
    fill_u = fill_util if datos['capital']['es_utilidad'] else fill_perd
    fnt_u  = fnt_util  if datos['capital']['es_utilidad'] else fnt_perd
    val_u  = float(datos['capital']['utilidad_ejercicio'])
    if not datos['capital']['es_utilidad']:
        val_u = -val_u
    cel(fila_pas, 4, lbl_util, fuente=fnt_u, aln=aln_left, relleno=fill_u)
    cel(fila_pas, 5, val_u, fuente=fnt_u, aln=aln_right, relleno=fill_u, fmt='#,##0.00')
    fila_pas += 1

    cel(fila_pas, 4, 'TOTAL CAPITAL', fuente=fnt_bold, aln=aln_right, relleno=fill_subtot2)
    cel(fila_pas, 5, float(datos['capital']['total']), fuente=fnt_bold, aln=aln_right, relleno=fill_subtot2, fmt='#,##0.00')
    fila_pas += 1

    cel(fila_pas, 4, 'TOTAL PASIVO + CAPITAL', fuente=fnt_white, aln=aln_right, relleno=fill_total)
    cel(fila_pas, 5, float(datos['total_pasivo_capital']), fuente=fnt_white, aln=aln_right, relleno=fill_total, fmt='#,##0.00')

    # Columna separadora vacía
    for r in range(fila_inicio, max(fila_act, fila_pas) + 1):
        ws.cell(row=r, column=3).border = borde

    ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    nombre = f"BalanceGeneral_{nombre_empresa.replace(' ','_')}_{fecha_hasta.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response


@login_required
def balance_general_pdf(request):
    from io import BytesIO
    from decimal import Decimal
    try:
        from xhtml2pdf import pisa
    except ImportError:
        messages.error(request, 'xhtml2pdf no está instalado.')
        return redirect('balance_general')

    empresa, periodo, fecha_hasta, datos = _get_balance_general_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    nombre_empresa = empresa.razon_social or empresa.nombre_comercial

    def filas_activo():
        rows = ''
        for grupo in datos['activo']['grupos']:
            rows += f'<tr class="sec-hdr"><td colspan="2">{grupo["nombre"].upper()}</td></tr>'
            for j, c in enumerate(grupo['cuentas']):
                cls = 'alt' if j % 2 else ''
                rows += f'<tr class="{cls}"><td class="ind">{c["cuenta"].nombre}</td><td class="num">{c["saldo"]:,.2f}</td></tr>'
            rows += f'<tr class="sub"><td>Total {grupo["nombre"]}</td><td class="num">{grupo["subtotal"]:,.2f}</td></tr>'
        rows += f'<tr class="tot"><td>TOTAL ACTIVO</td><td class="num">{datos["activo"]["total"]:,.2f}</td></tr>'
        return rows

    def filas_pasivo_capital():
        rows = '<tr class="col-hdr"><td colspan="2">PASIVO</td></tr>'
        for grupo in datos['pasivo']['grupos']:
            rows += f'<tr class="sec-hdr"><td colspan="2">{grupo["nombre"].upper()}</td></tr>'
            for j, c in enumerate(grupo['cuentas']):
                cls = 'alt' if j % 2 else ''
                rows += f'<tr class="{cls}"><td class="ind">{c["cuenta"].nombre}</td><td class="num">{c["saldo"]:,.2f}</td></tr>'
            rows += f'<tr class="sub"><td>Total {grupo["nombre"]}</td><td class="num">{grupo["subtotal"]:,.2f}</td></tr>'
        rows += f'<tr class="sub2"><td>TOTAL PASIVO</td><td class="num">{datos["pasivo"]["total"]:,.2f}</td></tr>'

        rows += '<tr class="col-hdr"><td colspan="2">CAPITAL</td></tr>'
        for grupo in datos['capital']['grupos']:
            rows += f'<tr class="sec-hdr"><td colspan="2">{grupo["nombre"].upper()}</td></tr>'
            for j, c in enumerate(grupo['cuentas']):
                cls = 'alt' if j % 2 else ''
                rows += f'<tr class="{cls}"><td class="ind">{c["cuenta"].nombre}</td><td class="num">{c["saldo"]:,.2f}</td></tr>'
            rows += f'<tr class="sub"><td>Total {grupo["nombre"]}</td><td class="num">{grupo["subtotal"]:,.2f}</td></tr>'

        lbl = 'Utilidad del Ejercicio' if datos['capital']['es_utilidad'] else 'Perdida del Ejercicio'
        cls_u = 'util' if datos['capital']['es_utilidad'] else 'perd'
        val_u = datos['capital']['utilidad_ejercicio']
        if not datos['capital']['es_utilidad']:
            rows += f'<tr class="{cls_u}"><td class="ind">{lbl}</td><td class="num">({val_u:,.2f})</td></tr>'
        else:
            rows += f'<tr class="{cls_u}"><td class="ind">{lbl}</td><td class="num">{val_u:,.2f}</td></tr>'

        rows += f'<tr class="sub2"><td>TOTAL CAPITAL</td><td class="num">{datos["capital"]["total"]:,.2f}</td></tr>'
        rows += f'<tr class="tot"><td>TOTAL PASIVO + CAPITAL</td><td class="num">{datos["total_pasivo_capital"]:,.2f}</td></tr>'
        return rows

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: letter portrait; margin: 1.5cm 1.8cm; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 8.5pt; color: #000; margin: 0; }}
  .enc-tabla {{ width: 100%; border-collapse: collapse; margin-bottom: 6pt; border: none; }}
  .enc-tabla td {{ border: none; padding: 0 2pt; background: white; }}
  .enc-titulo  {{ font-size: 13pt; font-weight: bold; text-align: center; margin: 0; }}
  .enc-empresa {{ font-size: 10pt; font-weight: bold; text-align: center; margin: 1pt 0 0 0; }}
  .enc-fechas  {{ font-size: 8.5pt; color: #333; text-align: center; margin: 1pt 0 0 0; }}
  .enc-moneda  {{ font-size: 8pt; color: #555; text-align: center; margin: 1pt 0 0 0; }}
  .columnas {{ width: 100%; border-collapse: collapse; }}
  .col-activo, .col-pasivo {{ width: 49%; vertical-align: top; padding: 0 3pt; }}
  .sep-col {{ width: 2%; }}
  table.detalle {{ width: 100%; border-collapse: collapse; font-size: 8pt; }}
  table.detalle td {{ border-left: 0.5pt solid #94a3b8; border-right: 0.5pt solid #94a3b8;
                      border-top: none; border-bottom: none; padding: 1px 4px; }}
  .enc-tabla td {{ border: none; padding: 0 2pt; background: white; }}
  .col-hdr td {{ background-color: #1B2B4E; color: white; font-weight: bold;
                  text-align: center; font-size: 8.5pt; padding: 3px;
                  border: 0.5pt solid #94a3b8; border-bottom: 1pt solid #94a3b8; }}
  .sec-hdr td {{ background-color: #D1D5DB; font-weight: bold; font-size: 8pt;
                  border: 0.5pt solid #94a3b8; border-bottom: 1pt solid #94a3b8; padding: 2px 5px; }}
  .alt td {{ background-color: #FAFAFA; }}
  .ind {{ padding-left: 12pt !important; }}
  .num {{ text-align: right; }}
  .sub  td {{ background-color: #EEF2F7; font-weight: bold;
              border-top: 0.8pt solid #000; border-bottom: none; }}
  .sub2 td {{ background-color: #D1D5DB; font-weight: bold;
              border-top: 1pt solid #000; border-bottom: 1pt solid #000; }}
  .tot  td {{ background-color: #1B2B4E; color: white; font-weight: bold;
              border-top: 1.5pt solid #000; border-bottom: 1.5pt solid #000; }}
  .util td {{ background-color: #D1FAE5; color: #065F46; font-weight: bold; }}
  .perd td {{ background-color: #FEE2E2; color: #991B1B; font-weight: bold; }}
</style></head><body>
<table class="enc-tabla">
  <tr>
    <td>
      <div class="enc-titulo">BALANCE GENERAL</div>
      <div class="enc-empresa">{nombre_empresa}</div>
      <div class="enc-fechas">Al {fecha_hasta.strftime('%d/%m/%Y')}</div>
      <div class="enc-moneda">(Expresado en Quetzales)</div>
    </td>
  </tr>
</table>
<table class="columnas">
<tr>
  <td class="col-activo">
    <table class="detalle">
      <tr class="col-hdr"><td colspan="2">ACTIVO</td></tr>
      {filas_activo()}
    </table>
  </td>
  <td class="sep-col"></td>
  <td class="col-pasivo">
    <table class="detalle">
      {filas_pasivo_capital()}
    </table>
  </td>
</tr>
</table>
</body></html>"""

    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
    if pisa_status.err:
        messages.error(request, 'Error al generar el PDF')
        return redirect('balance_general')

    buffer.seek(0)
    nombre = f"BalanceGeneral_{nombre_empresa.replace(' ','_')}_{fecha_hasta.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response


# ==================== BALANCE DE SALDOS ====================

def _get_balance_saldos_datos(request):
    from administracion.services.balance_saldos_service import BalanceSaldosService

    empresa_activa_id = request.session.get('empresa_activa_id')
    empresa_periodo_activo_id = request.session.get('empresa_periodo_activo_id')

    if not empresa_activa_id or not empresa_periodo_activo_id:
        return None, None, None, None, None

    try:
        empresa = Empresa.objects.get(id=empresa_activa_id)
        empresa_periodo = EmpresaPeriodo.objects.select_related('id_periodo').get(id=empresa_periodo_activo_id)
        periodo = empresa_periodo.id_periodo
    except (Empresa.DoesNotExist, EmpresaPeriodo.DoesNotExist):
        return None, None, None, None, None

    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    try:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date() if fecha_desde else periodo.fecha_inicial
    except ValueError:
        fecha_desde = periodo.fecha_inicial

    try:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date() if fecha_hasta else periodo.fecha_final
    except ValueError:
        fecha_hasta = periodo.fecha_final

    cuadros = BalanceSaldosService.get_datos_reporte(empresa_periodo, fecha_desde, fecha_hasta)
    return empresa, periodo, fecha_desde, fecha_hasta, cuadros


@login_required
def balance_saldos(request):
    empresa, periodo, fecha_desde, fecha_hasta, cuadros = _get_balance_saldos_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    if not cuadros:
        return render(request, 'administracion/reportes/balance_saldos.html', {
            'sin_datos': True,
            'empresa': empresa,
            'periodo': periodo,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        })

    return render(request, 'administracion/reportes/balance_saldos.html', {
        'sin_datos': False,
        'empresa': empresa,
        'periodo': periodo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'cuadros': cuadros,
        'fecha_reporte': timezone.now(),
    })


@login_required
def balance_saldos_excel(request):
    from decimal import Decimal

    empresa, periodo, fecha_desde, fecha_hasta, grupos, totales = _get_balance_saldos_datos(request)

    empresa, periodo, fecha_desde, fecha_hasta, cuadros = _get_balance_saldos_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    if not cuadros:
        messages.warning(request, 'No hay datos para exportar')
        return redirect('balance_saldos')

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance de Saldos"

    fnt_titulo = Font(bold=True, size=13)
    fnt_bold   = Font(bold=True, size=10)
    fnt_normal = Font(size=10)
    fnt_white  = Font(bold=True, size=10, color='FFFFFF')

    aln_center = Alignment(horizontal='center', vertical='center')
    aln_right  = Alignment(horizontal='right',  vertical='center')
    aln_left   = Alignment(horizontal='left',   vertical='center')

    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    fill_header = PatternFill(start_color='EEF2F7', end_color='EEF2F7', fill_type='solid')
    fill_mes    = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
    fill_sumas  = PatternFill(start_color='EEF2F7', end_color='EEF2F7', fill_type='solid')
    fill_alt    = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    nombre_empresa = empresa.razon_social or empresa.nombre_comercial
    fila = 1

    def merge(r, val, fuente, relleno, altura=14):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cl = ws.cell(row=r, column=1, value=val)
        cl.font = fuente; cl.alignment = aln_center; cl.fill = relleno
        for c in range(1, 4): ws.cell(row=r, column=c).border = borde
        ws.row_dimensions[r].height = altura

    def cel(r, c, val='', fuente=None, aln=None, relleno=None, fmt=None):
        cl = ws.cell(row=r, column=c, value=val)
        if fuente:  cl.font      = fuente
        if aln:     cl.alignment = aln
        if relleno: cl.fill      = relleno
        if fmt:     cl.number_format = fmt
        cl.border = borde
        return cl

    merge(fila, 'BALANCE DE SALDOS', fnt_titulo, PatternFill(fill_type=None), 18); fila += 1
    merge(fila, nombre_empresa, fnt_bold, PatternFill(fill_type=None)); fila += 1
    merge(fila,
        f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')} — (Expresado en Quetzales)",
        Font(size=9), PatternFill(fill_type=None), 12)
    fila += 2

    for cuadro in cuadros:
        # Encabezado mes
        merge(fila, cuadro['mes'], fnt_bold, fill_mes, 13); fila += 1

        # Encabezados columna
        cel(fila, 1, 'Nombre de la Cuenta', fuente=Font(bold=True, size=10), aln=aln_center, relleno=fill_header)
        cel(fila, 2, 'Debe',  fuente=Font(bold=True, size=10), aln=aln_center, relleno=fill_header)
        cel(fila, 3, 'Haber', fuente=Font(bold=True, size=10), aln=aln_center, relleno=fill_header)
        ws.row_dimensions[fila].height = 14; fila += 1

        for i, c in enumerate(cuadro['cuentas']):
            rf = fill_alt if i % 2 == 1 else None
            cel(fila, 1, c['nombre'],             fuente=fnt_normal, aln=aln_left,  relleno=rf)
            cel(fila, 2, float(c['debe'])  if c['debe']  > 0 else '', fuente=fnt_normal, aln=aln_right, relleno=rf, fmt='#,##0.00')
            cel(fila, 3, float(c['haber']) if c['haber'] > 0 else '', fuente=fnt_normal, aln=aln_right, relleno=rf, fmt='#,##0.00')
            fila += 1

        # Sumas iguales
        cel(fila, 1, 'Sumas Iguales', fuente=fnt_bold, aln=aln_right,  relleno=fill_sumas)
        cel(fila, 2, float(cuadro['total_debe']),  fuente=fnt_bold, aln=aln_right, relleno=fill_sumas, fmt='#,##0.00')
        cel(fila, 3, float(cuadro['total_haber']), fuente=fnt_bold, aln=aln_right, relleno=fill_sumas, fmt='#,##0.00')
        fila += 2

    ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    nombre = f"BalanceSaldos_{nombre_empresa.replace(' ','_')}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response


@login_required
def balance_saldos_pdf(request):
    from io import BytesIO
    from decimal import Decimal
    try:
        from xhtml2pdf import pisa
    except ImportError:
        messages.error(request, 'xhtml2pdf no está instalado.')
        return redirect('balance_saldos')

    empresa, periodo, fecha_desde, fecha_hasta, cuadros = _get_balance_saldos_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    if not cuadros:
        messages.warning(request, 'No hay datos para exportar')
        return redirect('balance_saldos')

    nombre_empresa = empresa.razon_social or empresa.nombre_comercial

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: letter portrait; margin: 1.5cm 1.8cm; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 8.5pt; color: #000; }}
  .enc-tabla {{ width: 100%; border-collapse: collapse; margin-bottom: 4pt; border: none; }}
  .enc-tabla td {{ border: none; padding: 0 2pt; background: white; }}
  .enc-izq {{ text-align: center; vertical-align: middle; width: 100%; padding: 0; }}
  .enc-titulo  {{ font-size: 13pt; font-weight: bold; text-align: center; }}
  .enc-empresa {{ font-size: 10pt; font-weight: bold; text-align: center; }}
  .enc-fechas  {{ font-size: 8.5pt; color: #333; text-align: center; }}
  .enc-moneda  {{ font-size: 8pt; color: #555; text-align: center; margin-bottom: 6pt; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 10pt; }}
  th {{ background-color: #EEF2F7; color: #0A1628; font-size: 8pt; padding: 2px 5px;
        border-left: 0.5pt solid #94a3b8; border-right: 0.5pt solid #94a3b8;
        border-top: 0.5pt solid #94a3b8; border-bottom: 1pt solid #94a3b8;
        text-align: center; font-weight: bold; }}
  td {{ border-left: 0.5pt solid #94a3b8; border-right: 0.5pt solid #94a3b8;
        border-top: none; border-bottom: none;
        padding: 1px 4px; vertical-align: middle; font-size: 8pt; }}
  .enc-tabla td {{ border: none; padding: 0 2pt; background: white; }}
  .col-cuenta {{ text-align: left; }}
  .col-monto  {{ width: 70pt; text-align: right; }}
  .row-mes td {{ background-color: #D1D5DB; color: #0A1628; font-weight: bold;
                 text-align: center; border: 0.5pt solid #94a3b8;
                 border-bottom: 1pt solid #94a3b8; padding: 2px 5px; }}
  .row-alt td  {{ background-color: #FAFAFA; }}
  .row-sumas td {{ background-color: #EEF2F7; font-weight: bold;
                   border-top: 1pt solid #000; border-bottom: 1pt solid #000; }}
</style></head><body>
<table class="enc-tabla">
  <tr>
    <td class="enc-izq">
      <div class="enc-titulo">BALANCE DE SALDOS</div>
      <div class="enc-empresa">{nombre_empresa}</div>
      <div class="enc-fechas">Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}</div>
      <div class="enc-moneda">(Expresado en Quetzales)</div>
    </td>
  </tr>
</table>
"""

    for cuadro in cuadros:
        html += f"""<table>
<tr class="row-mes"><td colspan="3">{cuadro['mes']}</td></tr>
<tr>
  <th class="col-cuenta">Nombre de la Cuenta</th>
  <th class="col-monto">Debe</th>
  <th class="col-monto">Haber</th>
</tr>
"""
        for j, c in enumerate(cuadro['cuentas']):
            cls  = 'row-alt' if j % 2 == 1 else ''
            debe  = f"{c['debe']:,.2f}"  if c['debe']  > 0 else ''
            haber = f"{c['haber']:,.2f}" if c['haber'] > 0 else ''
            html += f"""<tr class="{cls}">
  <td class="col-cuenta">{c['nombre']}</td>
  <td class="col-monto">{debe}</td>
  <td class="col-monto">{haber}</td>
</tr>"""

        html += f"""<tr class="row-sumas">
  <td class="col-cuenta" style="text-align:right;">Sumas Iguales</td>
  <td class="col-monto">{cuadro['total_debe']:,.2f}</td>
  <td class="col-monto">{cuadro['total_haber']:,.2f}</td>
</tr></table>"""

    html += '</body></html>'

    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
    if pisa_status.err:
        messages.error(request, 'Error al generar el PDF')
        return redirect('balance_saldos')

    buffer.seek(0)
    nombre = f"BalanceSaldos_{nombre_empresa.replace(' ','_')}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response


# ==================== LIBRO MAYOR ====================
def _get_libro_mayor_datos(request):
    from administracion.services.libro_mayor_service import LibroMayorService

    empresa_activa_id = request.session.get('empresa_activa_id')
    empresa_periodo_activo_id = request.session.get('empresa_periodo_activo_id')

    if not empresa_activa_id or not empresa_periodo_activo_id:
        return None, None, None, None, None, None

    try:
        empresa = Empresa.objects.get(id=empresa_activa_id)
        empresa_periodo = EmpresaPeriodo.objects.select_related('id_periodo').get(id=empresa_periodo_activo_id)
        periodo = empresa_periodo.id_periodo
    except (Empresa.DoesNotExist, EmpresaPeriodo.DoesNotExist):
        return None, None, None, None, None, None

    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    try:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date() if fecha_desde else periodo.fecha_inicial
    except ValueError:
        fecha_desde = periodo.fecha_inicial

    try:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date() if fecha_hasta else periodo.fecha_final
    except ValueError:
        fecha_hasta = periodo.fecha_final

    bloques = LibroMayorService.get_datos_reporte(empresa_periodo, fecha_desde, fecha_hasta)
    paginas = LibroMayorService.paginar(bloques)

    return empresa, periodo, fecha_desde, fecha_hasta, bloques, paginas


@login_required
def libro_mayor(request):
    empresa, periodo, fecha_desde, fecha_hasta, bloques, paginas = _get_libro_mayor_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    if not bloques:
        return render(request, 'administracion/reportes/libro_mayor.html', {
            'sin_datos': True,
            'empresa': empresa,
            'periodo': periodo,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        })

    return render(request, 'administracion/reportes/libro_mayor.html', {
        'sin_datos': False,
        'empresa': empresa,
        'periodo': periodo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'paginas': paginas,
        'total_cuentas': len(bloques),
        'fecha_reporte': timezone.now(),
    })


@login_required
def libro_mayor_excel(request):
    from decimal import Decimal

    empresa, periodo, fecha_desde, fecha_hasta, bloques, paginas = _get_libro_mayor_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    if not bloques:
        messages.warning(request, 'No hay datos para exportar')
        return redirect('libro_mayor')

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Mayor"

    fnt_titulo  = Font(bold=True, size=13)
    fnt_bold    = Font(bold=True, size=10)
    fnt_normal  = Font(size=10)
    fnt_italica = Font(size=9, italic=True, color='555555')

    aln_center = Alignment(horizontal='center', vertical='center')
    aln_right  = Alignment(horizontal='right',  vertical='center')
    aln_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    fill_cuenta  = PatternFill(start_color='1B2B4E', end_color='1B2B4E', fill_type='solid')
    fill_col_hdr = PatternFill(start_color='243B67', end_color='243B67', fill_type='solid')
    fill_total   = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    fill_ant     = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    fill_alt     = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

    fnt_white = Font(bold=True, size=10, color='FFFFFF')
    fnt_gold  = Font(bold=True, size=11, color='F4B41A')

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16

    fila = 1
    nombre_empresa = empresa.razon_social or empresa.nombre_comercial

    def merge_row(r, valor, fuente, relleno, altura=14):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cl = ws.cell(row=r, column=1, value=valor)
        cl.font = fuente; cl.alignment = aln_center; cl.fill = relleno
        for c in range(1, 8): ws.cell(row=r, column=c).border = borde
        ws.row_dimensions[r].height = altura

    def cel(r, c, valor='', fuente=None, aln=None, relleno=None, fmt=None):
        cl = ws.cell(row=r, column=c, value=valor)
        if fuente:  cl.font      = fuente
        if aln:     cl.alignment = aln
        if relleno: cl.fill      = relleno
        if fmt:     cl.number_format = fmt
        cl.border = borde
        return cl

    merge_row(fila, 'LIBRO MAYOR', fnt_titulo, PatternFill(fill_type=None), 18); fila += 1
    merge_row(fila, nombre_empresa, fnt_bold, PatternFill(fill_type=None)); fila += 1
    merge_row(fila,
        f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')} — "
        f"Período: {periodo.nombre} — (Expresado en Quetzales)",
        Font(size=9), PatternFill(fill_type=None), 12)
    fila += 2

    for bloque in bloques:
        cuenta = bloque['cuenta']
        nombre_cuenta = cuenta.nombre

        merge_row(fila, nombre_cuenta, fnt_gold, fill_cuenta, 15); fila += 1

        for c, txt in enumerate(['MES', 'DÍA', 'PARTIDA', 'NOMBRE', 'DEBE', 'HABER', 'SALDO'], 1):
            cel(fila, c, txt, fuente=fnt_white, aln=aln_center, relleno=fill_col_hdr)
        ws.row_dimensions[fila].height = 14; fila += 1

        if bloque['saldo_anterior'] != Decimal('0'):
            cel(fila, 1, '', relleno=fill_ant)
            cel(fila, 2, '', relleno=fill_ant)
            cel(fila, 3, '', relleno=fill_ant)
            cel(fila, 4, f"Saldo anterior al {fecha_desde.strftime('%d/%m/%Y')}", fuente=fnt_italica, aln=aln_left, relleno=fill_ant)
            cel(fila, 5, '', relleno=fill_ant)
            cel(fila, 6, '', relleno=fill_ant)
            cel(fila, 7, float(bloque['saldo_anterior']), fuente=fnt_bold, aln=aln_right, relleno=fill_ant, fmt='#,##0.00')
            fila += 1

        for i, f_mov in enumerate(bloque['filas']):
            rf = fill_alt if i % 2 == 1 else None
            cel(fila, 1, f_mov['texto_mes'],   fuente=fnt_normal, aln=aln_left,   relleno=rf)
            cel(fila, 2, f_mov['texto_dia'],   fuente=fnt_normal, aln=aln_center, relleno=rf)
            cel(fila, 3, f_mov['correlativo'], fuente=fnt_normal, aln=aln_center, relleno=rf)
            cel(fila, 4, f_mov['descripcion'], fuente=fnt_normal, aln=aln_left,   relleno=rf)
            cel(fila, 5, float(f_mov['debe'])  if f_mov['debe']  > 0 else '', fuente=fnt_normal, aln=aln_right, relleno=rf, fmt='#,##0.00')
            cel(fila, 6, float(f_mov['haber']) if f_mov['haber'] > 0 else '', fuente=fnt_normal, aln=aln_right, relleno=rf, fmt='#,##0.00')
            cel(fila, 7, float(f_mov['saldo']), fuente=fnt_bold, aln=aln_right, relleno=rf, fmt='#,##0.00')
            fila += 1

        cel(fila, 1, '', relleno=fill_total)
        cel(fila, 2, '', relleno=fill_total)
        cel(fila, 3, '', relleno=fill_total)
        cel(fila, 4, 'TOTALES', fuente=fnt_bold, aln=aln_right, relleno=fill_total)
        cel(fila, 5, float(bloque['total_debe']),  fuente=fnt_bold, aln=aln_right, relleno=fill_total, fmt='#,##0.00')
        cel(fila, 6, float(bloque['total_haber']), fuente=fnt_bold, aln=aln_right, relleno=fill_total, fmt='#,##0.00')
        cel(fila, 7, float(bloque['saldo_final']), fuente=fnt_bold, aln=aln_right, relleno=fill_total, fmt='#,##0.00')
        fila += 2

    ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    nombre = f"LibroMayor_{nombre_empresa.replace(' ','_')}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response


@login_required
def libro_mayor_pdf(request):
    from io import BytesIO
    from decimal import Decimal
    try:
        from xhtml2pdf import pisa
    except ImportError:
        messages.error(request, 'xhtml2pdf no está instalado.')
        return redirect('libro_mayor')

    empresa, periodo, fecha_desde, fecha_hasta, bloques, paginas = _get_libro_mayor_datos(request)

    if empresa is None:
        messages.warning(request, 'Debe seleccionar una empresa y un período activo')
        return redirect('home')

    if not bloques:
        messages.warning(request, 'No hay datos para exportar')
        return redirect('libro_mayor')

    nombre_empresa = empresa.razon_social or empresa.nombre_comercial

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: letter portrait; margin: 1.5cm 1.8cm; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 8.5pt; color: #000; }}
  .page-break {{ page-break-before: always; }}
  .enc-tabla {{ width: 100%; border-collapse: collapse; margin-bottom: 4pt; border: none; }}
  .enc-tabla td {{ border: none; padding: 0 2pt; background: white; }}
  .enc-izq {{ text-align: center; vertical-align: middle; width: 85%; padding: 0; }}
  .enc-der {{ text-align: right; vertical-align: top; padding: 0; font-size: 8.5pt; color: #333; white-space: nowrap; width: 15%; }}
  .enc-titulo  {{ font-size: 13pt; font-weight: bold; margin: 0; text-align: center; }}
  .enc-empresa {{ font-size: 10pt; font-weight: bold; margin: 1pt 0 0 0; text-align: center; }}
  .enc-fechas  {{ font-size: 8.5pt; color: #333; margin: 1pt 0 0 0; text-align: center; }}
  .enc-moneda  {{ font-size: 8pt; color: #555; margin: 1pt 0 0 0; text-align: center; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 8pt; }}
  th {{ background-color: #EEF2F7; color: #0A1628; font-size: 8pt; padding: 2px 4px;
        border-left: 0.5pt solid #94a3b8; border-right: 0.5pt solid #94a3b8;
        border-top: 0.5pt solid #94a3b8; border-bottom: 1pt solid #94a3b8;
        text-align: center; font-weight: bold; }}
  td {{ border-left: 0.5pt solid #94a3b8; border-right: 0.5pt solid #94a3b8;
        border-top: none; border-bottom: none;
        padding: 1px 4px; vertical-align: middle; font-size: 8pt; }}
  .enc-tabla td {{ border: none; padding: 0 2pt; background: white; }}
  .hdr-cuenta td {{ background-color: #D1D5DB; color: #0A1628; font-weight: bold;
                    font-size: 8.5pt; padding: 3px 6px;
                    border: 0.5pt solid #94a3b8; border-bottom: 1pt solid #94a3b8; }}
  .col-mes     {{ width: 42pt; text-align: left; }}
  .col-dia     {{ width: 18pt; text-align: center; }}
  .col-partida {{ width: 38pt; text-align: center; }}
  .col-desc    {{ text-align: left; }}
  .col-monto   {{ width: 62pt; text-align: right; }}
  .col-saldo   {{ width: 68pt; text-align: right; font-weight: bold; }}
  .row-ant td  {{ background-color: #EEF2F7; font-style: italic; color: #475569; }}
  .row-alt td  {{ background-color: #FAFAFA; }}
  .row-total td {{ background-color: #EEF2F7; font-weight: bold;
                   border-top: 1pt solid #000; border-bottom: 1pt solid #000; }}
  .row-total .col-saldo {{ color: #065f46; }}
</style></head><body>
"""

    for i, pagina in enumerate(paginas):
        if i > 0:
            html += '<div class="page-break"></div>'

        html += f"""
<table class="enc-tabla">
  <tr>
    <td class="enc-izq">
      <div class="enc-titulo">LIBRO MAYOR</div>
      <div class="enc-empresa">{nombre_empresa}</div>
      <div class="enc-fechas">Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}</div>
      <div class="enc-moneda">(Expresado en Quetzales)</div>
    </td>
    <td class="enc-der">Folio No. {pagina['numero']}</td>
  </tr>
</table>
"""
        for bloque in pagina['bloques']:
            cuenta = bloque['cuenta']
            nombre_cuenta = (
                f"{cuenta.id_area_contable.nombre} &rsaquo; "
                f"{cuenta.id_subgrupo.nombre} &rsaquo; "
                f"<strong>{cuenta.nombre}</strong>"
            )
            html += f"""<table>
<tr class="hdr-cuenta"><td colspan="7"><strong>{cuenta.nombre}</strong></td></tr>
<tr>
  <th class="col-mes">MES</th>
  <th class="col-dia">DÍA</th>
  <th class="col-partida">PARTIDA</th>
  <th class="col-desc">NOMBRE</th>
  <th class="col-monto">DEBE</th>
  <th class="col-monto">HABER</th>
  <th class="col-saldo">SALDO</th>
</tr>
"""
            if bloque['saldo_anterior'] != Decimal('0'):
                html += f"""<tr class="row-ant">
  <td class="col-mes"></td><td class="col-dia"></td><td class="col-partida"></td>
  <td class="col-desc">Saldo anterior al {fecha_desde.strftime('%d/%m/%Y')}</td>
  <td class="col-monto"></td><td class="col-monto"></td>
  <td class="col-saldo">{bloque['saldo_anterior']:,.2f}</td>
</tr>"""

            for j, fila in enumerate(bloque['filas']):
                cls   = 'row-alt' if j % 2 == 1 else ''
                debe  = f"{fila['debe']:,.2f}"  if fila['debe']  > 0 else ''
                haber = f"{fila['haber']:,.2f}" if fila['haber'] > 0 else ''
                html += f"""<tr class="{cls}">
  <td class="col-mes">{fila['texto_mes']}</td>
  <td class="col-dia">{fila['texto_dia']}</td>
  <td class="col-partida">{fila['correlativo']}</td>
  <td class="col-desc">{fila['descripcion']}</td>
  <td class="col-monto">{debe}</td>
  <td class="col-monto">{haber}</td>
  <td class="col-saldo">{fila['saldo']:,.2f}</td>
</tr>"""

            html += f"""<tr class="row-total">
  <td class="col-mes"></td><td class="col-dia"></td><td class="col-partida"></td>
  <td class="col-desc" style="text-align:right;">TOTALES</td>
  <td class="col-monto">{bloque['total_debe']:,.2f}</td>
  <td class="col-monto">{bloque['total_haber']:,.2f}</td>
  <td class="col-saldo">{bloque['saldo_final']:,.2f}</td>
</tr></table>"""

    html += '</body></html>'

    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
    if pisa_status.err:
        messages.error(request, 'Error al generar el PDF')
        return redirect('libro_mayor')

    buffer.seek(0)
    nombre = f"LibroMayor_{nombre_empresa.replace(' ','_')}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response



def migrar_catalogos_view(request):
    if not request.user.is_superuser:
        return render(request, 'administracion/migrar_catalogos.html', {
            'error': 'Solo los administradores pueden ejecutar esta acción.',
            'ejecutado': False,
        })

    if request.method == 'POST':
        from io import StringIO, TextIOWrapper
        from django.core.management import call_command
        accion = request.POST.get('accion', 'catalogos')
        output = StringIO()
        error = None

        try:
            if accion == 'asientos':
                f_asiento   = request.FILES.get('csv_asiento')
                f_movimiento = request.FILES.get('csv_movimiento')
                f_detalle   = request.FILES.get('csv_detalle')

                if not f_asiento or not f_movimiento or not f_detalle:
                    error = 'Debe subir los tres archivos CSV para migrar asientos.'
                else:
                    call_command(
                        'migrar_asientos',
                        stdout=output,
                        stderr=output,
                        asiento_file=f_asiento,
                        movimiento_file=f_movimiento,
                        detalle_file=f_detalle,
                    )
            else:
                archivos_requeridos = [
                    'AreaContable.csv', 'Grupo.csv', 'SubGrupo.csv', 'Cuenta.csv',
                    'Periodo.csv', 'Empresa.csv', 'EmpresaPeriodo.csv',
                    'Sucursal.csv', 'Proveedor.csv',
                ]
                archivos = {}
                faltantes = []
                for nombre in archivos_requeridos:
                    clave = 'csv_' + nombre.replace('.csv', '').lower()
                    f = request.FILES.get(clave)
                    if f:
                        archivos[nombre] = f
                    else:
                        faltantes.append(nombre)

                if faltantes:
                    error = f'Archivos faltantes: {", ".join(faltantes)}'
                else:
                    call_command(
                        'migrar_catalogos',
                        stdout=output, stderr=output,
                        files=archivos,
                    )
        except Exception as e:
            error = str(e)

        log = output.getvalue()
        return render(request, 'administracion/migrar_catalogos.html', {
            'ejecutado': True,
            'accion': accion,
            'log': log,
            'error': error,
        })

    return render(request, 'administracion/migrar_catalogos.html', {'ejecutado': False})



# ==================== API SELECTOR RÁPIDO ====================

@login_required
def api_empresas_lista(request):
    """Retorna todas las empresas para el selector rápido."""
    empresas = Empresa.objects.order_by('nombre_comercial', 'razon_social').values(
        'id', 'nombre_comercial', 'razon_social', 'nit'
    )
    data = [
        {
            'id': e['id'],
            'nombre': e['nombre_comercial'] or e['razon_social'],
            'nit': e['nit'],
        }
        for e in empresas
    ]
    return JsonResponse({'empresas': data})


@login_required
def api_empresa_periodos(request, empresa_id):
    """Retorna los períodos asignados a una empresa para el selector rápido."""
    try:
        empresa = Empresa.objects.get(id=empresa_id)
    except Empresa.DoesNotExist:
        return JsonResponse({'periodos': []})

    eps = (
        EmpresaPeriodo.objects
        .filter(id_empresa=empresa)
        .select_related('id_periodo')
        .order_by('-id_periodo__fecha_inicial')
    )

    periodos = [
        {
            'id':     ep.id,
            'nombre': ep.id_periodo.nombre,
            'activo': bool(ep.estatus),
        }
        for ep in eps
    ]
    return JsonResponse({'periodos': periodos})